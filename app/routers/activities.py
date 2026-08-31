from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import or_, and_
from sqlalchemy.orm import Session, selectinload
from sqlalchemy.orm.attributes import flag_modified

from auth import get_current_user, log_action, require_permission, user_has_permission
from calculators.baseline_updater import update_baseline_from_activity, is_auto_approval_enabled
from calculators.pay_period import get_billing_period, get_or_create_period_for_date
from database.models import Activity, ActivitySource, ActivityStatus, AppUser, Employee, EmployeeSpringerFlag, PayPeriod, PayPeriodStatus
from database.schemas import (
    ActivityApprove,
    ActivityCreate,
    ActivityDeactivate,
    ActivityResponse,
    ActivitySplit,
    ActivityUpdate,
    VagtplanHideBody,
)
from database.session import get_db

router = APIRouter(prefix="/api/activities", tags=["activities"])

_toggle_springer_access = require_permission("toggle_springer")


class SpringerFlagUpdate(BaseModel):
    employee_id: int
    pay_period_id: int
    enabled: bool

_XLSX = Path(__file__).parent.parent / "Fraværstyper.xlsx"
_LABEL_OVERRIDES = {"Kursus/Skole": "skole_kursus"}


def _normalize_type(label: str) -> str:
    if label in _LABEL_OVERRIDES:
        return _LABEL_OVERRIDES[label]
    s = label.lower()
    s = s.replace("§", "paragraf_")
    s = s.replace("æ", "ae").replace("ø", "oe").replace("å", "aa")
    s = s.replace(" ", "_").replace("/", "_").replace(".", "").replace("-", "_")
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")


_BACKEND_ONLY_TYPES = {
    "sygdom_u_8uger",       # gammel form (evt. eksisterende DB-rækker)
    "sygdom_u_8_uger",      # normaliseret form fra Fraværstyper.xlsx
    "barn_1sygedag_u_8uger",
    "barsel_u_loen",
}  # tildeles automatisk, ikke brugervalgt

_HIDDEN_FROM_TYPE_PICKER = {
    "dob_overnatning",  # tilgås kun via DOB-krydsfeltet i Overnatning-modalen, ikke egen type
}

_NINE_MONTHS = 9


def _months_between(d1: date, d2: date) -> int:
    """Antal hele måneder fra d1 til d2."""
    months = (d2.year - d1.year) * 12 + (d2.month - d1.month)
    if d2.day < d1.day:
        months -= 1
    return months


def _has_vagtplan_edit_access(db: Session, current_user: AppUser, emp: Employee) -> bool:
    """Kun relevant når en aktivitet oprettes med source='vagtplan' (fra Vagtplan-griddet).
    Almindelig oprettelse/redigering fra Aktivitetsoversigten er upåvirket – der er i dag
    ingen rolle-baseret restriktion på selve /api/activities uden for dette."""
    if user_has_permission(db, current_user, "vagtplan_edit_all"):
        return True
    if user_has_permission(db, current_user, "vagtplan_edit_own"):
        return bool(emp.initials) and emp.initials.strip().lower() == current_user.initials.strip().lower()
    return False


@router.get("/absence-types")
def get_absence_types(
    current_user: AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from database.models import MasterAbsenceType, MasterPayType
    rows = db.query(MasterAbsenceType).filter(
        MasterAbsenceType.is_active == True
    ).order_by(MasterAbsenceType.sort_order, MasterAbsenceType.label).all()
    if rows:
        result = [{"value": r.normalized_key, "label": r.label} for r in rows]
    else:
        # fallback til Excel
        wb = load_workbook(_XLSX, read_only=True, data_only=True)
        ws = wb.active
        result = []
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            label = row[0]
            if label:
                value = _normalize_type(str(label).strip())
                if value in _BACKEND_ONLY_TYPES:
                    continue
                result.append({"value": value, "label": str(label).strip()})
        wb.close()
    # Tilføj brugerdefinerede løntypekoder som aktivitetstyper (spring over dubletter)
    existing_keys = {r["value"] for r in result}
    user_pay_types = db.query(MasterPayType).filter(
        MasterPayType.is_user_created == True
    ).order_by(MasterPayType.sort_order).all()
    result.extend([
        {"value": upt.code_key, "label": upt.label}
        for upt in user_pay_types
        if upt.code_key not in existing_keys and upt.code_key not in _HIDDEN_FROM_TYPE_PICKER
    ])
    return result


FOUR_HOURS = 4 * 60
TWELVE_HOURS = 12 * 60


def _duration_minutes(a: Activity) -> int:
    total = int((a.end_time - a.start_time).total_seconds() // 60)
    # Segmenter er autoritativ kilde for DDD-aktiviteter (pause_intervals kan være forældet
    # efter 'Ret linje'/'Tilpas'). Manuelle aktiviteter har ingen segmenter → brug pause_intervals.
    if a.segments:
        pauses = [[seg[0], seg[1]] for seg in a.segments if len(seg) >= 3 and seg[2] == "rest"]
    else:
        pauses = list(a.pause_intervals or [])
    for p in pauses:
        try:
            ps = datetime.fromisoformat(p[0])
            pe = datetime.fromisoformat(p[1])
            actual_start = max(a.start_time, ps)
            actual_end = min(a.end_time, pe)
            if actual_end > actual_start:
                total -= int((actual_end - actual_start).total_seconds() // 60)
        except (ValueError, IndexError):
            pass
    return max(0, total)


def _recalculate_pcts(a: Activity) -> None:
    """Genberegn procentfordelinger ud fra det aktuelle segments-felt."""
    segments = a.segments or []
    total_sec = (a.end_time - a.start_time).total_seconds()
    if total_sec <= 0:
        return
    type_sec: dict[str, float] = {"driving": 0.0, "work": 0.0, "availability": 0.0, "rest": 0.0}
    for seg in segments:
        try:
            s = datetime.fromisoformat(seg[0])
            e = datetime.fromisoformat(seg[1])
            t = seg[2]
            if t in type_sec:
                type_sec[t] += (e - s).total_seconds()
        except (ValueError, IndexError):
            pass
    a.driving_pct           = Decimal(str(round(type_sec["driving"]      / total_sec * 100, 2)))
    a.other_work_pct        = Decimal(str(round(type_sec["work"]         / total_sec * 100, 2)))
    a.availability_time_pct = Decimal(str(round(type_sec["availability"] / total_sec * 100, 2)))
    a.rest_pause_pct        = Decimal(str(round(type_sec["rest"]         / total_sec * 100, 2)))


class SegmentCorrectionBody(BaseModel):
    segment_index: int
    revert: bool = False


class SegmentResizeBody(BaseModel):
    segment_index: int
    new_end_iso: str  # "YYYY-MM-DDTHH:MM" eller "YYYY-MM-DDTHH:MM:SS"


def _duration_minutes_in_window(a: Activity, window_start: datetime, window_end: datetime) -> int:
    """Antal arbejdsminutter (pauser fratrukket) af aktiviteten der overlapper [window_start, window_end)."""
    start = max(a.start_time, window_start)
    end = min(a.end_time, window_end)
    if end <= start:
        return 0
    total = int((end - start).total_seconds() // 60)
    if a.segments:
        pauses = [[seg[0], seg[1]] for seg in a.segments if len(seg) >= 3 and seg[2] == "rest"]
    else:
        pauses = list(a.pause_intervals or [])
    for p in pauses:
        try:
            ps = datetime.fromisoformat(p[0])
            pe = datetime.fromisoformat(p[1])
            actual_start = max(start, ps)
            actual_end = min(end, pe)
            if actual_end > actual_start:
                total -= int((actual_end - actual_start).total_seconds() // 60)
        except (ValueError, IndexError):
            pass
    return max(0, total)


def _day_reaches_4h_with_approved(a: Activity, dur: int) -> bool:
    """True hvis denne aktivitets varighed sammen med andre godkendte aktiviteters
    overlap med samme kalenderdag for medarbejderen når op på mindst 4 timer.
    En natvagt der starter dagen før men fortsætter ind på denne dag tæller
    kun med de minutter der reelt ligger på denne kalenderdag."""
    db = Session.object_session(a)
    if db is None:
        return False
    day_start = datetime.combine(a.start_time.date(), datetime.min.time())
    day_end = day_start + timedelta(days=1)
    others = (
        db.query(Activity)
        .filter(
            Activity.employee_id == a.employee_id,
            Activity.id != a.id,
            Activity.status == ActivityStatus.approved,
            Activity.start_time < day_end,
            Activity.end_time > day_start,
        )
        .all()
    )
    total = dur + sum(_duration_minutes_in_window(o, day_start, day_end) for o in others)
    return total >= FOUR_HOURS


def _to_response(a: Activity) -> ActivityResponse:
    dur = _duration_minutes(a)
    under_4h = dur < FOUR_HOURS and not _day_reaches_4h_with_approved(a, dur)
    return ActivityResponse(
        id=a.id,
        employee_id=a.employee_id,
        employee_name=a.employee.name,
        employee_number=a.employee.employee_number,
        pay_period_id=a.pay_period_id,
        source=a.source,
        activity_type=a.activity_type,
        start_time=a.start_time,
        end_time=a.end_time,
        duration_minutes=dur,
        availability_time_pct=a.availability_time_pct,
        rest_pause_pct=a.rest_pause_pct,
        other_work_pct=a.other_work_pct,
        driving_pct=a.driving_pct,
        loading_minutes=a.loading_minutes,
        unloading_minutes=a.unloading_minutes,
        pause_intervals=a.pause_intervals or [],
        segments=a.segments or [],
        is_edited=a.original_start_time is not None,
        has_split_children=len(a.split_children) > 0,
        parent_activity_id=a.parent_activity_id,
        status=a.status,
        approved_by=a.approved_by,
        approved_at=a.approved_at,
        deactivated_by=a.deactivated_by,
        comment=a.comment,
        is_under_4h=under_4h,
        is_over_12h=dur > TWELVE_HOURS,
        is_manual=a.source == ActivitySource.manual,
        created_by=a.created_by,
        vehicle_registration=a.vehicle_registration,
        vehicle_number=a.vehicle_number,
        km_start=a.km_start,
        km_end=a.km_end,
        salt_supplement=bool(a.salt_supplement),
        auto_approved=bool(a.auto_approved),
        auto_approval_flags=a.auto_approval_flags or [],
        is_likely_incomplete=bool(a.is_likely_incomplete),
        hidden_from_vagtplan=bool(a.hidden_from_vagtplan),
    )


@router.get("", response_model=list[ActivityResponse])
def list_activities(
    period_start: Optional[str] = None,
    employee_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return activities either for a pay period (period_start, default = today's period)
    or for an arbitrary date range (date_from/date_to) independent of pay periods –
    den sidste bruges af Vagtplan, hvis 4-ugers vindue ikke følger 14-dages lønperioder."""
    if date_from and date_to:
        range_start = datetime.combine(date.fromisoformat(date_from), datetime.min.time())
        range_end = datetime.combine(date.fromisoformat(date_to), datetime.min.time()) + timedelta(days=1)
        q = (
            db.query(Activity)
            .join(Activity.employee)
            .options(
                selectinload(Activity.employee),
                selectinload(Activity.split_children),
            )
            .filter(Activity.start_time < range_end, Activity.end_time > range_start)
        )
        if employee_id:
            q = q.filter(Activity.employee_id == employee_id)
        activities = q.order_by(Activity.start_time).all()
        return [_to_response(a) for a in activities]

    start_date = date.fromisoformat(period_start) if period_start else date.today()
    period = get_or_create_period_for_date(start_date, db)

    period_start_dt = datetime.combine(period.start_date, datetime.min.time())

    q = (
        db.query(Activity)
        .join(Activity.employee)
        .options(
            selectinload(Activity.employee),
            selectinload(Activity.split_children),
        )
        .filter(
            or_(
                Activity.pay_period_id == period.id,
                # Aktiviteter fra forrige periode der slutter i denne (krydser periodegraensen)
                and_(
                    Activity.end_time >= period_start_dt,
                    Activity.start_time < period_start_dt,
                ),
            )
        )
    )
    if employee_id:
        q = q.filter(Activity.employee_id == employee_id)

    activities = q.order_by(Activity.start_time).all()
    return [_to_response(a) for a in activities]


@router.get("/period-info")
def period_info(period_start: Optional[str] = None,
                current_user: AppUser = Depends(get_current_user),
                db: Session = Depends(get_db)):
    """Return current period stats and prev/next periods."""
    from datetime import date, timedelta
    from database.models import PayPeriod

    start_date = date.fromisoformat(period_start) if period_start else date.today()
    period = get_or_create_period_for_date(start_date, db)

    activities = db.query(Activity).filter(Activity.pay_period_id == period.id).all()
    counts = {"pending": 0, "approved": 0, "deactivated": 0}
    for a in activities:
        counts[a.status.value] += 1

    # Faste 14-dages perioder: forrige/næste kan altid beregnes
    prev_start = (period.start_date - timedelta(days=14)).isoformat()
    next_start = (period.start_date + timedelta(days=14)).isoformat()

    return {
        "period": {
            "id": period.id,
            "start_date": period.start_date.isoformat(),
            "end_date": period.end_date.isoformat(),
            "status": period.status.value,
            "total": len(activities),
            "pending": counts["pending"],
            "approved": counts["approved"],
            "deactivated": counts["deactivated"],
        },
        "prev_period_start": prev_start,
        "next_period_start": next_start,
    }


@router.get("/springer-flags")
def get_springer_flags(pay_period_id: int,
                        current_user: AppUser = Depends(get_current_user),
                        db: Session = Depends(get_db)):
    rows = db.query(EmployeeSpringerFlag).filter(
        EmployeeSpringerFlag.pay_period_id == pay_period_id,
        EmployeeSpringerFlag.enabled == True,
    ).all()
    return {r.employee_id: True for r in rows}


@router.post("/springer-flag")
def set_springer_flag(body: SpringerFlagUpdate,
                       current_user: AppUser = Depends(_toggle_springer_access),
                       db: Session = Depends(get_db)):
    period = db.query(PayPeriod).filter(PayPeriod.id == body.pay_period_id).first()
    if not period:
        raise HTTPException(404, "Lønperiode ikke fundet")
    if period.status == PayPeriodStatus.closed:
        raise HTTPException(400, "Lønperioden er låst – kan ikke ændres")
    row = db.query(EmployeeSpringerFlag).filter(
        EmployeeSpringerFlag.employee_id == body.employee_id,
        EmployeeSpringerFlag.pay_period_id == body.pay_period_id,
    ).first()
    if row:
        row.enabled = body.enabled
        row.updated_by = current_user.initials
    else:
        row = EmployeeSpringerFlag(
            employee_id=body.employee_id, pay_period_id=body.pay_period_id,
            enabled=body.enabled, updated_by=current_user.initials,
        )
        db.add(row)
    db.commit()
    log_action(db, current_user, "springer_flag_set", "employee_springer_flag", body.employee_id,
               f"periode {body.pay_period_id}: {'sat' if body.enabled else 'fjernet'}")
    db.commit()
    return {"employee_id": body.employee_id, "pay_period_id": body.pay_period_id, "enabled": body.enabled}


_EIGHT_WEEKS = 56  # dage


@router.post("", response_model=ActivityResponse, status_code=201)
def create_manual_activity(body: ActivityCreate,
                            current_user: AppUser = Depends(get_current_user),
                            db: Session = Depends(get_db)):
    emp = db.query(Employee).filter(Employee.id == body.employee_id).first()
    if not emp:
        raise HTTPException(404, "Medarbejder ikke fundet")

    activity_source = ActivitySource.manual
    if body.source == "vagtplan":
        if not _has_vagtplan_edit_access(db, current_user, emp):
            raise HTTPException(403, "Ingen redigeringsret til Vagtplan for denne medarbejder")
        activity_source = ActivitySource.vagtplan

    activity_type = body.activity_type

    if activity_type in _BACKEND_ONLY_TYPES:
        raise HTTPException(400, "Denne aktivitetstype tildeles automatisk og kan ikke angives manuelt")

    # Sygdom: 8 uger eller under → uden løn
    if activity_type == "sygdom":
        employed_days = (body.start_time.date() - emp.hire_date).days
        if employed_days <= _EIGHT_WEEKS:
            activity_type = "sygdom_u_8uger"

    # Barn 1.sygedag: under 9 måneder → dagpengesats
    if activity_type == "barn_1sygedag":
        if _months_between(emp.hire_date, body.start_time.date()) < _NINE_MONTHS:
            activity_type = "barn_1sygedag_u_8uger"

    # Anciennitetskontrol for barsel: terminsdato skal være min. 9 måneder efter ansættelse
    if activity_type == "barsel":
        if body.terminsdato is None:
            raise HTTPException(400, "Terminsdato er påkrævet for barsel")
        if _months_between(emp.hire_date, body.terminsdato) < _NINE_MONTHS:
            activity_type = "barsel_u_loen"
        # Husk seneste terminsdato på medarbejderen, så den foreslås ved næste barsel-oprettelse
        emp.terminsdato = body.terminsdato

    period = get_billing_period(body.start_time.date(), db)
    is_absence = activity_type != "normal"
    can_auto_approve = (
        user_has_permission(db, current_user, "auto_approve_manual_activities")
        and is_auto_approval_enabled(db)
    )
    activity = Activity(
        employee_id=body.employee_id,
        pay_period_id=period.id,
        source=activity_source,
        created_by=current_user.initials,
        activity_type=activity_type,
        start_time=body.start_time,
        end_time=body.end_time,
        loading_minutes=body.loading_minutes,
        unloading_minutes=body.unloading_minutes,
        comment=body.comment,
        vehicle_number=body.vehicle_number,
        km_start=body.km_start,
        km_end=body.km_end,
        salt_supplement=body.salt_supplement,
        pause_intervals=body.pause_intervals,
        status=ActivityStatus.pending,
    )
    db.add(activity)
    db.flush()

    if is_absence or can_auto_approve:
        activity.status = ActivityStatus.approved
        activity.approved_by = current_user.initials
        activity.approved_at = datetime.utcnow()

        if can_auto_approve and not activity.comment:
            dur = _duration_minutes(activity)
            if dur < FOUR_HOURS and not _day_reaches_4h_with_approved(activity, dur):
                activity.comment = current_user.initials

    log_action(db, current_user, "create_activity", "activity", activity.id,
               f"Manuelt oprettet for {emp.name}")
    db.commit()
    db.refresh(activity)
    return _to_response(activity)


@router.get("/{activity_id}", response_model=ActivityResponse)
def get_activity(activity_id: int,
                 current_user: AppUser = Depends(get_current_user),
                 db: Session = Depends(get_db)):
    a = db.query(Activity).filter(Activity.id == activity_id).first()
    if not a:
        raise HTTPException(404, "Aktivitet ikke fundet")
    return _to_response(a)


@router.patch("/{activity_id}", response_model=ActivityResponse)
def update_activity(activity_id: int, body: ActivityUpdate,
                    current_user: AppUser = Depends(get_current_user),
                    db: Session = Depends(get_db)):
    a = db.query(Activity).filter(Activity.id == activity_id).first()
    if not a:
        raise HTTPException(404, "Aktivitet ikke fundet")
    if body.activity_type and body.activity_type in _BACKEND_ONLY_TYPES:
        raise HTTPException(400, "Denne aktivitetstype tildeles automatisk og kan ikke angives manuelt")
    # Gem originale tider ved første rettelse (muliggør fortryd)
    times_changed = (
        (body.start_time and body.start_time != a.start_time)
        or (body.end_time and body.end_time != a.end_time)
    )
    if times_changed and a.original_start_time is None:
        a.original_start_time = a.start_time
        a.original_end_time = a.end_time
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(a, field, value)
    if body.pause_intervals is not None:
        flag_modified(a, "pause_intervals")
        for p in a.pause_intervals:
            try:
                ps = datetime.fromisoformat(p[0])
                pe = datetime.fromisoformat(p[1])
                if ps < a.start_time:
                    raise HTTPException(
                        400, f"Pause ({ps.strftime('%H:%M')}) starter før aktiviteten begynder ({a.start_time.strftime('%H:%M')})"
                    )
                if pe > a.end_time:
                    raise HTTPException(
                        400, f"Pause ({pe.strftime('%H:%M')}) slutter efter aktiviteten er slut ({a.end_time.strftime('%H:%M')})"
                    )
            except HTTPException:
                raise
            except (ValueError, IndexError):
                pass
    # Update pay period if start_time changed
    if body.start_time:
        period = get_billing_period(body.start_time.date(), db)
        a.pay_period_id = period.id
    log_action(db, current_user, "update_activity", "activity", a.id)
    db.commit()
    db.refresh(a)
    return _to_response(a)


@router.post("/{activity_id}/undo-edit", response_model=ActivityResponse)
def undo_edit(activity_id: int,
              current_user: AppUser = Depends(get_current_user),
              db: Session = Depends(get_db)):
    """Fortryd manuelle tidsændringer – gendan de originale tider."""
    a = db.query(Activity).filter(Activity.id == activity_id).first()
    if not a:
        raise HTTPException(404, "Aktivitet ikke fundet")
    if a.original_start_time is None:
        raise HTTPException(400, "Ingen tidsændringer at fortryde")
    a.start_time = a.original_start_time
    a.end_time = a.original_end_time
    a.original_start_time = None
    a.original_end_time = None
    period = get_or_create_period_for_date(a.start_time.date(), db)
    a.pay_period_id = period.id
    db.commit()
    db.refresh(a)
    return _to_response(a)


@router.post("/{activity_id}/undo-split", response_model=ActivityResponse)
def undo_split(activity_id: int,
               current_user: AppUser = Depends(get_current_user),
               db: Session = Depends(get_db)):
    """
    Fortryd et split. Kan kaldes på originalen eller en af delene.
    Delene slettes, og den originale aktivitet sættes tilbage til afventende.
    """
    a = db.query(Activity).filter(Activity.id == activity_id).first()
    if not a:
        raise HTTPException(404, "Aktivitet ikke fundet")
    # Find originalen (forælderen) hvis der klikkes på en del
    parent = a if a.parent_activity_id is None else (
        db.query(Activity).filter(Activity.id == a.parent_activity_id).first()
    )
    if parent is None:
        raise HTTPException(404, "Original aktivitet ikke fundet")
    children = db.query(Activity).filter(Activity.parent_activity_id == parent.id).all()
    if not children:
        raise HTTPException(400, "Aktiviteten er ikke splittet")
    for child in children:
        db.delete(child)
    parent.status = ActivityStatus.pending
    parent.approved_by = None
    parent.approved_at = None
    parent.deactivated_by = None
    db.commit()
    db.refresh(parent)
    return _to_response(parent)


@router.post("/{activity_id}/approve", response_model=ActivityResponse)
def approve_activity(activity_id: int, body: ActivityApprove,
                     current_user: AppUser = Depends(get_current_user),
                     db: Session = Depends(get_db)):
    a = db.query(Activity).filter(Activity.id == activity_id).first()
    if not a:
        raise HTTPException(404, "Aktivitet ikke fundet")
    if a.status == ActivityStatus.deactivated:
        raise HTTPException(400, "Kan ikke godkende en deaktiveret aktivitet")
    dur = _duration_minutes(a)
    if dur < FOUR_HOURS and not _day_reaches_4h_with_approved(a, dur) and not body.comment:
        raise HTTPException(
            400,
            "Aktivitet er under 4 timer – angiv en begrundelse i kommentarfeltet"
        )
    a.status = ActivityStatus.approved
    a.approved_by = current_user.initials
    a.approved_at = datetime.utcnow()
    if body.comment:
        a.comment = body.comment
    log_action(db, current_user, "approve", "activity", a.id,
               f"Godkendt for {a.employee.name} ({a.start_time.strftime('%d-%m-%Y')})")
    db.commit()
    update_baseline_from_activity(a, db)
    db.commit()
    db.refresh(a)
    return _to_response(a)


@router.post("/auto-approve-pending")
def bulk_auto_approve(
    period_start: Optional[str] = None,
    current_user: AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Auto-godkend alle egnede pending-aktiviteter i en lønperiode."""
    from datetime import date as _date
    from datetime import datetime as _dt
    from calculators.auto_approval import should_auto_approve

    if not is_auto_approval_enabled(db):
        raise HTTPException(400, "Automatisk godkendelse er slået fra i systemindstillinger")

    start_date = _date.fromisoformat(period_start) if period_start else _date.today()
    period = get_or_create_period_for_date(start_date, db)

    pending = (
        db.query(Activity)
        .filter(
            Activity.pay_period_id == period.id,
            Activity.status == ActivityStatus.pending,
        )
        .all()
    )

    approved_count = 0
    flagged_count = 0

    for act in pending:
        ok, flags = should_auto_approve(act, db)
        if ok:
            act.status = ActivityStatus.approved
            act.auto_approved = True
            act.auto_approval_flags = []
            act.approved_by = "AUTO"
            act.approved_at = _dt.utcnow()
            db.commit()
            update_baseline_from_activity(act, db)
            approved_count += 1
        else:
            act.auto_approval_flags = flags
            db.commit()
            flagged_count += 1

    log_action(db, current_user, "auto_approve_bulk", details=f"periode={period.start_date}, godkendt={approved_count}, flagget={flagged_count}")
    db.commit()
    return {"approved": approved_count, "flagged": flagged_count}


@router.post("/{activity_id}/deactivate", response_model=ActivityResponse)
def deactivate_activity(activity_id: int, body: ActivityDeactivate,
                        current_user: AppUser = Depends(get_current_user),
                        db: Session = Depends(get_db)):
    a = db.query(Activity).filter(Activity.id == activity_id).first()
    if not a:
        raise HTTPException(404, "Aktivitet ikke fundet")
    a.status = ActivityStatus.deactivated
    a.deactivated_by = current_user.initials
    a.approved_at = datetime.utcnow()
    if body.comment is not None:
        a.comment = body.comment
    log_action(db, current_user, "deactivate", "activity", a.id,
               f"Deaktiveret for {a.employee.name} ({a.start_time.strftime('%d-%m-%Y')})")
    db.commit()
    db.refresh(a)
    return _to_response(a)


@router.post("/{activity_id}/hide-from-vagtplan", response_model=ActivityResponse)
def hide_from_vagtplan(activity_id: int, body: VagtplanHideBody,
                       current_user: AppUser = Depends(get_current_user),
                       db: Session = Depends(get_db)):
    a = db.query(Activity).filter(Activity.id == activity_id).first()
    if not a:
        raise HTTPException(404, "Aktivitet ikke fundet")
    a.hidden_from_vagtplan = body.hidden
    log_action(db, current_user, "hide_from_vagtplan", "activity", a.id,
               f"{'Skjult' if body.hidden else 'Vist'} i Vagtplan for {a.employee.name}")
    db.commit()
    db.refresh(a)
    return _to_response(a)


@router.post("/{activity_id}/correct-segment", response_model=ActivityResponse)
def correct_segment(
    activity_id: int,
    body: SegmentCorrectionBody,
    current_user: AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Ret et segment (rest→work) eller gendan det. Gemmer original type som 4. element."""
    a = (
        db.query(Activity)
        .options(selectinload(Activity.employee), selectinload(Activity.split_children))
        .filter(Activity.id == activity_id)
        .first()
    )
    if not a:
        raise HTTPException(404, "Aktivitet ikke fundet")

    segments = [list(seg) for seg in (a.segments or [])]
    idx = body.segment_index
    if idx < 0 or idx >= len(segments):
        raise HTTPException(400, "Ugyldigt segmentindeks")

    seg = segments[idx]
    if body.revert:
        if len(seg) < 4:
            raise HTTPException(400, "Segment er ikke rettet")
        seg = seg[:2] + [seg[3]]          # gendan original type
    else:
        if seg[2] != "rest":
            raise HTTPException(400, "Kun pausesegmenter kan rettes")
        if len(seg) >= 4:
            raise HTTPException(400, "Segment er allerede rettet")
        seg = seg[:2] + ["work", seg[2]]  # ret til arbejde, bevar original

    segments[idx] = seg
    a.segments = segments
    flag_modified(a, "segments")
    _recalculate_pcts(a)
    db.commit()
    db.refresh(a)
    log_action(db, current_user, "correct_segment", "activity", activity_id,
               {"segment_index": idx, "revert": body.revert})
    return _to_response(a)


def _correct_all_segments_list(segments: list) -> tuple[list, int]:
    """Ret alle u-rettede pause-segmenter ('rest') til 'work', bevarer original
    type som 4. element (samme transformation som correct_segment, uden revert).
    Returnerer (nye segmenter, antal rettede linjer)."""
    result = []
    corrected = 0
    for seg in segments:
        seg = list(seg)
        if seg[2] == "rest" and len(seg) < 4:
            seg = seg[:2] + ["work", seg[2]]
            corrected += 1
        result.append(seg)
    return result, corrected


@router.post("/{activity_id}/correct-all-segments", response_model=ActivityResponse)
def correct_all_segments(
    activity_id: int,
    current_user: AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Ret alle u-rettede pause-segmenter for aktiviteten til 'Andet arbejde' i én omgang."""
    a = (
        db.query(Activity)
        .options(selectinload(Activity.employee), selectinload(Activity.split_children))
        .filter(Activity.id == activity_id)
        .first()
    )
    if not a:
        raise HTTPException(404, "Aktivitet ikke fundet")

    new_segments, corrected_count = _correct_all_segments_list(a.segments or [])
    if corrected_count == 0:
        raise HTTPException(400, "Ingen pauselinjer at rette")

    a.segments = new_segments
    flag_modified(a, "segments")
    _recalculate_pcts(a)
    db.commit()
    db.refresh(a)
    log_action(db, current_user, "correct_all_segments", "activity", activity_id,
               {"corrected_count": corrected_count})
    return _to_response(a)


@router.post("/{activity_id}/resize-segment", response_model=ActivityResponse)
def resize_segment(
    activity_id: int,
    body: SegmentResizeBody,
    current_user: AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Tilpas længden af et pausesegment.
    Kortere: det overskydende tid tilføjes som nyt 'work'-segment.
    Længere: næste segment forkortes tilsvarende.
    """
    from datetime import datetime as _dt

    a = (
        db.query(Activity)
        .options(selectinload(Activity.employee), selectinload(Activity.split_children))
        .filter(Activity.id == activity_id)
        .first()
    )
    if not a:
        raise HTTPException(404, "Aktivitet ikke fundet")

    segments = [list(seg) for seg in (a.segments or [])]
    idx = body.segment_index
    if idx < 0 or idx >= len(segments):
        raise HTTPException(400, "Ugyldigt segmentindeks")

    seg = segments[idx]
    if seg[2] != "rest":
        raise HTTPException(400, "Kun pausesegmenter kan tilpasses")

    new_end_str = body.new_end_iso if len(body.new_end_iso) > 16 else body.new_end_iso + ":00"
    try:
        new_end_dt = _dt.fromisoformat(new_end_str)
    except ValueError:
        raise HTTPException(400, "Ugyldigt tidspunkt")

    seg_start_dt = _dt.fromisoformat(seg[0])
    seg_end_dt   = _dt.fromisoformat(seg[1])

    if new_end_dt <= seg_start_dt:
        raise HTTPException(400, "Ny sluttid skal være efter segmentets starttid")
    if new_end_dt == seg_end_dt:
        raise HTTPException(400, "Ingen ændring")

    if new_end_dt < seg_end_dt:
        # Forkortelse: del resten op som nyt arbejdssegment
        shortened = seg[:2] + [seg[2]] + (seg[3:] if len(seg) > 3 else [])
        shortened[1] = new_end_str
        new_work = [new_end_str, seg[1], "work"]
        segments = segments[:idx] + [shortened, new_work] + segments[idx + 1:]
    else:
        # Forlængelse: lån tid fra næste segment
        if idx + 1 >= len(segments):
            raise HTTPException(400, "Der er intet næste segment at tage tid fra")
        next_seg = segments[idx + 1]
        next_end_dt = _dt.fromisoformat(next_seg[1])
        if new_end_dt >= next_end_dt:
            raise HTTPException(400, "Ny sluttid overstiger næste segments sluttid")
        extended = seg[:2] + [seg[2]] + (seg[3:] if len(seg) > 3 else [])
        extended[1] = new_end_str
        shortened_next = [new_end_str] + next_seg[1:]
        segments = segments[:idx] + [extended, shortened_next] + segments[idx + 2:]

    a.segments = segments
    flag_modified(a, "segments")
    _recalculate_pcts(a)
    a.is_edited = True
    db.commit()
    db.refresh(a)
    log_action(db, current_user, "resize_segment", "activity", activity_id,
               {"segment_index": idx, "new_end_iso": body.new_end_iso})
    return _to_response(a)


@router.post("/{activity_id}/reopen", response_model=ActivityResponse)
def reopen_activity(activity_id: int,
                    current_user: AppUser = Depends(get_current_user),
                    db: Session = Depends(get_db)):
    a = db.query(Activity).filter(Activity.id == activity_id).first()
    if not a:
        raise HTTPException(404, "Aktivitet ikke fundet")
    a.status = ActivityStatus.pending
    a.approved_by = None
    a.approved_at = None
    a.hidden_from_vagtplan = False
    log_action(db, current_user, "reopen_activity", "activity", a.id,
               f"Genåbnet for {a.employee.name} ({a.start_time.strftime('%d-%m-%Y')})")
    db.commit()
    db.refresh(a)
    return _to_response(a)


@router.post("/{activity_id}/split", response_model=list[ActivityResponse])
def split_activity(activity_id: int, body: ActivitySplit,
                   current_user: AppUser = Depends(get_current_user),
                   db: Session = Depends(get_db)):
    """
    Split an activity at body.split_at.
    Both parts: pending.
    """
    a = db.query(Activity).filter(Activity.id == activity_id).first()
    if not a:
        raise HTTPException(404, "Aktivitet ikke fundet")
    if not (a.start_time < body.split_at < a.end_time):
        raise HTTPException(400, "Splitpunkt skal ligge mellem start- og sluttid")

    # Mark original as replaced (track who split it)
    a.status = ActivityStatus.deactivated
    a.deactivated_by = current_user.initials
    a.approved_at = datetime.utcnow()

    # Fordel pauseintervaller til de to dele (klippes ved splitpunktet)
    pauses1, pauses2 = [], []
    for s, e in (a.pause_intervals or []):
        p_start, p_end = datetime.fromisoformat(s), datetime.fromisoformat(e)
        if p_start < body.split_at:
            pauses1.append([p_start.isoformat(), min(p_end, body.split_at).isoformat()])
        if p_end > body.split_at:
            pauses2.append([max(p_start, body.split_at).isoformat(), p_end.isoformat()])

    # Fordel hændelsessegmenter ligeså
    segs1, segs2 = [], []
    for s, e, name in (a.segments or []):
        s_start, s_end = datetime.fromisoformat(s), datetime.fromisoformat(e)
        if s_start < body.split_at:
            segs1.append([s_start.isoformat(), min(s_end, body.split_at).isoformat(), name])
        if s_end > body.split_at:
            segs2.append([max(s_start, body.split_at).isoformat(), s_end.isoformat(), name])

    # Fælles felter kopieres til begge dele
    common = dict(
        employee_id=a.employee_id,
        pay_period_id=a.pay_period_id,
        source=a.source,
        activity_type=a.activity_type,
        loading_minutes=a.loading_minutes,
        unloading_minutes=a.unloading_minutes,
        salt_supplement=a.salt_supplement,
        vehicle_registration=a.vehicle_registration,
        vehicle_number=a.vehicle_number,
        km_start=a.km_start,
        km_end=a.km_end,
        status=ActivityStatus.pending,
        parent_activity_id=a.id,
    )

    part1 = Activity(
        **common,
        start_time=a.start_time,
        end_time=body.split_at,
        pause_intervals=pauses1,
        segments=segs1,
        split_part=1,
        comment="Split: første del",
    )
    part2 = Activity(
        **common,
        start_time=body.split_at,
        end_time=a.end_time,
        pause_intervals=pauses2,
        segments=segs2,
        split_part=2,
        comment="Split: anden del",
    )
    _recalculate_pcts(part1)
    _recalculate_pcts(part2)
    db.add(part1)
    db.add(part2)
    db.flush()
    log_action(db, current_user, "split", "activity", a.id,
               f"Splittet for {a.employee.name} ({a.start_time.strftime('%d-%m-%Y')})")
    db.commit()
    db.refresh(part1)
    db.refresh(part2)
    return [_to_response(part1), _to_response(part2)]
