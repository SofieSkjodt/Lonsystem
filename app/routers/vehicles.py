from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session

from auth import get_current_user, require_permission
from database.session import get_db
from database.models import Activity, AppUser, Vehicle
from database.schemas import VehicleCreate, VehicleUpdate, VehicleResponse

router = APIRouter(prefix="/api/vehicles", tags=["vehicles"])

XLSX_PATH = Path(__file__).resolve().parent.parent / "Vognnumre.xlsx"


@router.get("", response_model=list[VehicleResponse])
def list_vehicles(current_user: AppUser = Depends(get_current_user),
                  db: Session = Depends(get_db)):
    return db.query(Vehicle).order_by(Vehicle.registration_number).all()


@router.post("", response_model=VehicleResponse, status_code=201)
def create_vehicle(body: VehicleCreate,
                   current_user: AppUser = Depends(require_permission("manage_vehicles")),
                   db: Session = Depends(get_db)):
    reg = body.registration_number.strip()
    if db.query(Vehicle).filter(Vehicle.registration_number == reg).first():
        raise HTTPException(400, "Registreringsnummer eksisterer allerede")
    v = Vehicle(registration_number=reg, vehicle_number=body.vehicle_number.strip())
    db.add(v)
    db.commit()
    db.refresh(v)
    return v


@router.patch("/{vehicle_id}", response_model=VehicleResponse)
def update_vehicle(vehicle_id: int, body: VehicleUpdate,
                   current_user: AppUser = Depends(require_permission("manage_vehicles")),
                   db: Session = Depends(get_db)):
    v = db.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
    if not v:
        raise HTTPException(404, "Vogn ikke fundet")
    if body.registration_number is not None:
        reg = body.registration_number.strip()
        existing = db.query(Vehicle).filter(
            Vehicle.registration_number == reg,
            Vehicle.id != vehicle_id,
        ).first()
        if existing:
            raise HTTPException(400, "Registreringsnummer eksisterer allerede")
        v.registration_number = reg
    if body.vehicle_number is not None:
        v.vehicle_number = body.vehicle_number.strip()
    db.commit()
    db.refresh(v)
    return v


@router.delete("/{vehicle_id}", status_code=204)
def delete_vehicle(vehicle_id: int,
                   current_user: AppUser = Depends(require_permission("manage_vehicles")),
                   db: Session = Depends(get_db)):
    v = db.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
    if not v:
        raise HTTPException(404, "Vogn ikke fundet")
    if db.query(Activity).filter(Activity.vehicle_registration == v.registration_number).count() > 0:
        raise HTTPException(400, "Vognen er tilknyttet aktiviteter og kan ikke slettes")
    db.delete(v)
    db.commit()


@router.post("/import-xlsx")
def import_from_xlsx(current_user: AppUser = Depends(get_current_user),
                     db: Session = Depends(get_db)):
    """Importer vogne fra app/Vognnumre.xlsx."""
    if not XLSX_PATH.exists():
        raise HTTPException(404, f"Filen '{XLSX_PATH.name}' blev ikke fundet i app-mappen")

    from openpyxl import load_workbook
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active

    imported = 0
    skipped = 0
    seen = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        reg = str(row[0]).strip()
        vnum = str(row[1]).strip()
        if not reg or not vnum:
            continue
        if reg.lower() in ("registreringsnummer", "reg.nr", "regnr"):
            continue

        if reg in seen:
            skipped += 1
            continue
        seen.add(reg)

        if db.query(Vehicle).filter(Vehicle.registration_number == reg).first():
            skipped += 1
            continue

        db.add(Vehicle(registration_number=reg, vehicle_number=vnum))
        imported += 1

    db.commit()
    return {"imported": imported, "skipped": skipped}
