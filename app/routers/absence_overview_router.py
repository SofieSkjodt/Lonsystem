from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path
from typing import Optional
import io

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from auth import get_current_user, require_permission
from calculators.pay_period import get_or_create_period_for_date
from calculators.rates_loader import load_agreement_types_from_db
from database.models import Activity, ActivityStatus, AppUser, Employee
from database.session import get_db

router = APIRouter(prefix="/api/absence-overview", tags=["absence-overview"])

_access = require_permission("absence_overview")

_XLSX = Path(__file__).parent.parent / "Fraværstyper.xlsx"

_FALLBACK_LABELS = {
    "normal":       "Normal tid",
    "ferie":        "Ferie",
    "fri":          "Fri",
    "afspadsering": "Afspadsering",
    "skole_kursus": "Kursus/Skole",
}


def _normalize_type(label: str) -> str:
    _OVERRIDES = {"Kursus/Skole": "skole_kursus"}
    if label in _OVERRIDES:
        return _OVERRIDES[label]
    s = label.lower()
    s = s.replace("§", "paragraf_")
    s = s.replace("æ", "ae").replace("ø", "oe").replace("å", "aa")
    s = s.replace(" ", "_").replace("/", "_").replace(".", "").replace("-", "_")
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")


def _load_type_labels(db: Session) -> dict:
    """Returns {value: label} from MasterAbsenceType DB + fallbacks."""
    from database.models import MasterAbsenceType
    labels = dict(_FALLBACK_LABELS)
    try:
        rows = db.query(MasterAbsenceType).all()
        if rows:
            for r in rows:
                labels[r.normalized_key] = r.label
            return labels
    except Exception:
        pass
    # fallback til Excel hvis DB er tom
    try:
        wb = load_workbook(_XLSX, read_only=True, data_only=True)
        ws = wb.active
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            cell = row[0]
            if cell:
                label = str(cell).strip()
                labels[_normalize_type(label)] = label
        wb.close()
    except Exception:
        pass
    return labels


def _resolve_dates(date_from: Optional[str], date_to: Optional[str], db: Session):
    if date_from and date_to:
        return date.fromisoformat(date_from), date.fromisoformat(date_to)
    period = get_or_create_period_for_date(date.today(), db)
    return period.start_date, period.end_date


_PAID_ABSENCE_TYPES = {"sygdom", "barn_1sygedag", "barn_2_3sygedag", "barsel", "graviditetsbetinget_sygdom", "skole_kursus"}  # timesats = medarbejderens overenskomstsats
_FIXED_RATE_ABSENCE = {"paragraf_56_syg": 137.43}                    # fast sats uanset overenskomsttype


def _compute_data(d_from: date, d_to: date, db: Session) -> dict:
    type_labels = _load_type_labels(db)
    agreement_rates = load_agreement_types_from_db(db)

    activities = (
        db.query(Activity)
        .join(Activity.employee)
        .filter(
            Activity.status == ActivityStatus.approved,
            Activity.activity_type != "normal",
            Activity.start_time >= d_from.isoformat(),
            Activity.start_time < (d_to + timedelta(days=1)).isoformat(),
            Employee.active == True,
        )
        .order_by(Employee.first_name, Employee.last_name, Activity.start_time)
        .all()
    )

    all_types: set = set()
    emp_map: dict = {}

    for act in activities:
        emp = act.employee
        if emp.id not in emp_map:
            hourly_rate = float(agreement_rates.get(emp.agreement_type, 0))
            emp_map[emp.id] = {
                "employee_id":     emp.id,
                "employee_name":   emp.name,
                "employee_number": emp.employee_number,
                "hourly_rate":     hourly_rate,
                "absences":        defaultdict(lambda: {"minutes": 0, "day_set": set()}),
            }
        atype = act.activity_type
        all_types.add(atype)
        duration_min = int((act.end_time - act.start_time).total_seconds() // 60)
        emp_map[emp.id]["absences"][atype]["minutes"] += duration_min

        cur_day = act.start_time.date()
        end_day = act.end_time.date()
        while cur_day <= end_day:
            emp_map[emp.id]["absences"][atype]["day_set"].add(cur_day)
            cur_day += timedelta(days=1)

    employees = []
    for emp_info in emp_map.values():
        absences = {}
        for atype, data in emp_info["absences"].items():
            label = type_labels.get(atype, atype.replace("_", " ").capitalize())
            if atype in _FIXED_RATE_ABSENCE:
                rate = _FIXED_RATE_ABSENCE[atype]
            elif atype in _PAID_ABSENCE_TYPES:
                rate = emp_info["hourly_rate"]
            else:
                rate = 0.0
            absences[atype] = {
                "label": label,
                "hours": round(data["minutes"] / 60, 2),
                "days":  len(data["day_set"]),
                "rate":  rate,
            }
        employees.append({
            "employee_id":     emp_info["employee_id"],
            "employee_name":   emp_info["employee_name"],
            "employee_number": emp_info["employee_number"],
            "absences":        absences,
        })

    sorted_types = sorted(all_types, key=lambda t: type_labels.get(t, t))

    return {
        "date_from":     d_from.isoformat(),
        "date_to":       d_to.isoformat(),
        "absence_types": [{"value": t, "label": type_labels.get(t, t)} for t in sorted_types],
        "employees":     employees,
    }


@router.get("/data")
def absence_data(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: AppUser = Depends(_access),
    db: Session = Depends(get_db),
):
    d_from, d_to = _resolve_dates(date_from, date_to, db)
    return _compute_data(d_from, d_to, db)


@router.get("/employee-options")
def employee_options(
    current_user: AppUser = Depends(_access),
    db: Session = Depends(get_db),
):
    emps = (
        db.query(Employee)
        .filter(Employee.active == True)
        .order_by(Employee.first_name, Employee.last_name)
        .all()
    )
    groups = sorted({e.dispatcher_group for e in emps if e.dispatcher_group})
    return {
        "employees": [{"id": e.id, "name": e.name, "dispatcher_group": e.dispatcher_group} for e in emps],
        "dispatcher_groups": groups,
    }


@router.get("/export-per-employee")
def export_per_employee(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    employee_id: Optional[int] = None,
    dispatcher_group: Optional[str] = None,
    current_user: AppUser = Depends(_access),
    db: Session = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    d_from, d_to = _resolve_dates(date_from, date_to, db)
    data = _compute_data(d_from, d_to, db)

    # Filter employees
    employees = data["employees"]
    if employee_id:
        employees = [e for e in employees if e["employee_id"] == employee_id]
    elif dispatcher_group:
        # Look up which employee ids belong to this dispatcher group
        group_emp_ids = {
            e.id for e in db.query(Employee).filter(
                Employee.active == True,
                Employee.dispatcher_group == dispatcher_group,
            ).all()
        }
        employees = [e for e in employees if e["employee_id"] in group_emp_ids]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fravær per medarbejder"

    header_fill = PatternFill(start_color="317423", end_color="317423", fill_type="solid")
    row_fill    = PatternFill(start_color="D4EDCC", end_color="D4EDCC", fill_type="solid")

    ws.append(["Medarbejder", "Lønnr", "Fraværstype", "Dage", "Timer", "Sats"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    ws.freeze_panes = "A2"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12

    sorted_emps = sorted(employees, key=lambda e: e["employee_name"])
    for emp_idx, emp in enumerate(sorted_emps):
        sorted_abs = sorted(emp["absences"].items(), key=lambda x: x[1]["label"])
        use_fill = (emp_idx % 2 == 0)
        for i, (_, ainfo) in enumerate(sorted_abs):
            rate_val = ainfo.get("rate", 0)
            ws.append([
                emp["employee_name"] if i == 0 else "",
                emp["employee_number"] if i == 0 else "",
                ainfo["label"],
                ainfo["days"],
                ainfo["hours"],
                rate_val if rate_val else "",
            ])
            if use_fill:
                for cell in ws[ws.max_row]:
                    cell.fill = row_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    def _safe(s):
        return s.replace(" ", "_").replace("/", "_")

    if employee_id and employees:
        suffix = f"_{_safe(employees[0]['employee_name'])}"
    elif dispatcher_group:
        suffix = f"_{_safe(dispatcher_group)}"
    else:
        suffix = ""
    filename = f"fravaer_per_medarbejder{suffix}_{d_from}_{d_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export-per-type")
def export_per_type(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    absence_type: Optional[str] = None,
    current_user: AppUser = Depends(_access),
    db: Session = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    d_from, d_to = _resolve_dates(date_from, date_to, db)
    data = _compute_data(d_from, d_to, db)

    # Aggregate totals per type across all employees (optionally filtered)
    type_totals: dict = {}
    type_labels: dict = {}
    for emp in data["employees"]:
        for atype, ainfo in emp["absences"].items():
            if absence_type and atype != absence_type:
                continue
            if atype not in type_totals:
                type_totals[atype] = {"days": 0, "hours": 0.0}
                type_labels[atype] = ainfo["label"]
            type_totals[atype]["days"]  += ainfo["days"]
            type_totals[atype]["hours"]  = round(type_totals[atype]["hours"] + ainfo["hours"], 2)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fravær per type"

    header_fill = PatternFill(start_color="317423", end_color="317423", fill_type="solid")
    row_fill    = PatternFill(start_color="D4EDCC", end_color="D4EDCC", fill_type="solid")

    ws.append(["Fraværstype", "Dage i alt", "Timer i alt"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    ws.freeze_panes = "A2"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14

    sorted_types = sorted(type_totals.items(), key=lambda x: type_labels.get(x[0], x[0]))
    for i, (atype, totals) in enumerate(sorted_types):
        ws.append([type_labels.get(atype, atype), totals["days"], totals["hours"]])
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = row_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    type_label_clean = type_labels.get(absence_type, absence_type).replace(" ", "_").replace("/", "_") if absence_type else ""
    type_suffix = f"_{type_label_clean}" if type_label_clean else ""
    filename = f"fravaer_per_type{type_suffix}_{d_from}_{d_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
