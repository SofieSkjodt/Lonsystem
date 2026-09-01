"""
Lønkørsel:
- /api/payroll/preview           – JSON-mellemregninger til forsiden
- /api/payroll/proevekoersel     – Excel-fil (alle eller én medarbejder)
- /api/payroll/export-csv        – "Kør løn": Danløn CSV
- /api/payroll/pdf-timesedler    – dan PDF-timesedler pr. medarbejder (gemmes
                                   lokalt i output/timesedler; e-mail-afsendelse
                                   tilføjes senere, jf. aftale 10/6-2026)
"""
import csv
import io
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from auth import get_current_user, log_action, require_permission
from database.models import AppUser
from utils.safe_paths import is_under_allowed_root

from calculators.overtime import (
    OT_13_KEY,
    OT_13_MAX,
    OT_BEFORE_KEY,
    OT_EXTRA_KEY,
    calculate_overtime,
    calculate_flat_hours,
)
from calculators.day_type import (
    DayType,
    classify_day,
    compute_sh_hours,
    calculate_special_day_overtime,
)
from calculators.pay_period import get_or_create_period_for_date, is_even_week
import logging as _logging
from calculators.rates_loader import (
    load_agreement_types_from_db,
    load_overtime_rates_from_db,
    load_salt_supplement_rate_from_db,
    load_overnight_rate_from_db,
    load_dob_overnight_rate_from_db,
    load_dagpenge_rate_from_db,
    load_springer_rate_from_db,
    load_overtime_rates_by_id_from_db,
    load_supplement_rates_by_id_from_db,
    get_active_supplement_for_period,
)
from database.models import Activity, ActivityStatus, AgreementKind, Employee, EmployeeSpringerFlag, Holiday, MasterCvrNumber, PayPeriod, PayPeriodStatus
from database.session import get_db

router = APIRouter(prefix="/api/payroll", tags=["payroll"])

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "output"

_payroll_access = require_permission("payroll")
_reopen_access = require_permission("reopen_period")


def _safe_save_dir(raw_path: str) -> Path:
    """Valider at den bruger-angivne gemmesti er tilladt: ikke inde i selve
    applikationsmappen, og under en af de tilladte rødder (se utils/safe_paths.py)
    – forhindrer at CSV/PDF-eksport skrives til en vilkårlig placering på
    serverens diske."""
    save_dir = Path(raw_path).resolve()
    app_dir = BASE_DIR.resolve()
    try:
        save_dir.relative_to(app_dir)
        raise HTTPException(400, "Ugyldig gemmemappe – kan ikke gemme i applikationsmappen")
    except ValueError:
        pass  # Stien er IKKE inde i app-mappen – det er ønsket
    if not is_under_allowed_root(save_dir):
        raise HTTPException(400, "Ugyldig gemmemappe – skal ligge under din hjemmemappe")
    return save_dir

TWO = Decimal("0.01")


def _round2(v: Decimal) -> Decimal:
    return v.quantize(TWO, rounding=ROUND_HALF_UP)


def _get_employee_cvr(emp: Employee, db: Session) -> str:
    """Returnerer CVR-nummeret for medarbejderen (eget hvis sat, ellers standard fra Stamdata)."""
    if emp.cvr_number:
        return emp.cvr_number
    default = db.query(MasterCvrNumber).filter(MasterCvrNumber.is_default == True).first()
    if default:
        return default.cvr_number
    raise HTTPException(500, "Intet standard-CVR fundet – tilføj et CVR-nummer i Stamdata → CVR-nummer")


def _get_pay_type_data(db: Session) -> dict:
    """Returnerer {code_key: {code, in_csv, qty_type, rate_src, inc_total}} fra stamdata-tabellen."""
    from database.models import MasterPayType
    rows = db.query(MasterPayType).all()
    return {r.code_key.upper(): {
        "code": r.danloen_code,
        "in_csv": r.include_in_csv,
        "qty_type": r.csv_quantity_type or "hours",
        "rate_src": r.csv_rate_source or "hourly",
        "inc_rate": r.csv_include_rate if r.csv_include_rate is not None else True,
        "inc_total": r.csv_include_total or False,
    } for r in rows}


def _resolve_rate(rate_src: str, calc: dict) -> float:
    """Opslår en sats fra calc-dict baseret på rate_src-streng."""
    if rate_src.startswith("overtime:"):
        rid = int(rate_src.split(":", 1)[1])
        return float(calc.get("ot_rates_by_id", {}).get(rid, 0))
    if rate_src.startswith("supplement:"):
        rid = int(rate_src.split(":", 1)[1])
        return float(calc.get("supplement_rates_by_id", {}).get(rid, 0))
    # Gamle faste værdier – bevaret som sikkerhedsnet efter migrationen til id-baserede referencer.
    if rate_src == "ot_before":
        return float(calc["ot_rates"][OT_BEFORE_KEY])
    if rate_src == "ot_13":
        return float(calc["ot_rates"][OT_13_KEY])
    if rate_src == "ot_extra":
        return float(calc["ot_rates"][OT_EXTRA_KEY])
    if rate_src == "salt":
        return float(calc.get("salt_rate", 0))
    if rate_src == "overnight":
        return float(calc.get("overnight_rate", 0))
    if rate_src == "dagpenge":
        return float(calc.get("dagpenge_sats", 137.43))
    if rate_src == "springer":
        return float(calc.get("springer_rate", 0))
    return float(calc["hourly_rate"])


def _user_pay_type_rows(emp_id: int, start: date, end: date, calc: dict, db: Session) -> list:
    """Beregner antal/timer pr. brugerdefineret løntypekode for én medarbejder i perioden."""
    from database.models import MasterPayType
    user_types = db.query(MasterPayType).filter(MasterPayType.is_user_created == True).all()
    result = []
    start_str = start.isoformat()
    end_str = (end + timedelta(days=1)).isoformat()
    for upt in user_types:
        acts = db.query(Activity).filter(
            Activity.employee_id == emp_id,
            Activity.activity_type == upt.code_key,
            Activity.status == ActivityStatus.approved,
            Activity.start_time >= start_str,
            Activity.start_time < end_str,
        ).all()
        qty_type = upt.csv_quantity_type or "hours"
        if qty_type == "count":
            qty = float(len(acts))
        else:
            qty = sum((a.end_time - a.start_time).total_seconds() / 3600 for a in acts)
        if qty > 0:
            rate = _resolve_rate(upt.csv_rate_source or "hourly", calc)
            result.append((upt.code_key.upper(), qty, rate))
    return result


def _springer_row(calc: dict) -> tuple:
    """CSV-tuple for springertillæg: samme timetal som løntypekode 1 (Normal tid),
    men kun hvis flaget er sat for medarbejderens periode (calc['springer_enabled'])."""
    qty = calc["normal_hours"] if calc.get("springer_enabled") else 0
    return ("SPRINGERTILLAEG", qty, calc.get("springer_rate", 0))


def _builtin_absence_qty(pt: dict, key: str, activity_type: str, hours_value: float,
                          emp_id: int, start: date, end: date, db: Session) -> float:
    """Antal for en indbygget fraværs-løntype – bruger stamdatas csv_quantity_type
    (fx 'count' for antal dage) i stedet for altid at bruge det akkumulerede
    timeantal fra _calculate_employee."""
    qty_type = pt.get(key, {}).get("qty_type", "hours")
    if qty_type != "count":
        return hours_value
    start_str = start.isoformat()
    end_str = (end + timedelta(days=1)).isoformat()
    return float(db.query(Activity).filter(
        Activity.employee_id == emp_id,
        Activity.activity_type == activity_type,
        Activity.status == ActivityStatus.approved,
        Activity.start_time >= start_str,
        Activity.start_time < end_str,
    ).count())


def _normal_hours_for_day(emp: Employee, d: date) -> Decimal:
    """Normaltid for en given dag fra medarbejderens timefordeling."""
    schedule = emp.work_schedule or {"even": [0] * 7, "odd": [0] * 7}
    key = "even" if is_even_week(d) else "odd"
    return Decimal(str(schedule[key][d.weekday()]))


def _afspadsering_hours(emp: Employee, act: Activity) -> Decimal:
    """Timer for en afspadsering-aktivitet. En periode (aktiviteten strækker sig
    over flere kalenderdage, jf. 'Til dato' i oprettelsesmodalen) tæller 7,4 t
    (eller medarbejderens skemalagte timer) pr. hverdag i perioden – uanset de
    faktiske klokketider på aktiviteten. En enkeltdags-aktivitet bruger stadig
    den reelle varighed, så en delvis fridag kan registreres korrekt."""
    start_d = act.start_time.date()
    end_d = act.end_time.date()
    if start_d == end_d:
        return Decimal(str((act.end_time - act.start_time).total_seconds())) / 3600
    total = Decimal("0")
    cur_d = start_d
    while cur_d <= end_d:
        if cur_d.weekday() < 5:
            day_h = _normal_hours_for_day(emp, cur_d)
            total += day_h if day_h > 0 else Decimal("7.4")
        cur_d += timedelta(days=1)
    return total


@dataclass
class _DayPiece:
    """Et stykke af en aktivitet der falder inden for én kalenderdag."""
    start_time: datetime
    end_time: datetime
    pause_intervals: list
    activity_type: str
    salt_supplement: bool
    vehicle_number: Optional[str]


def _split_into_day_pieces(act: Activity) -> list[_DayPiece]:
    """Splitter en aktivitet, der strækker sig over midnat, i ét stykke pr.
    kalenderdag – hvert stykke skal beregnes under SIN EGEN dags dag-type og
    normaltids-loft (fx en søndag-til-mandag-vagt: søndagsdelen får
    søndagsregler, mandagsdelen får mandagens normale tidsvindues-beregning).
    Pauseintervaller der overlapper døgnskellet klippes tilsvarende.

    Kaldes KUN når aktiviteten starter på en søndag/helligdag (se
    _calculate_employee) – normaltids-/OT13-loftet hører til vagten og skal
    IKKE nulstilles ved midnat for almindelige hverdage/lørdage (bekræftet af
    bruger 2026-07-02).
    """
    if act.segments:
        raw_pauses = [
            (datetime.fromisoformat(seg[0]), datetime.fromisoformat(seg[1]))
            for seg in act.segments if len(seg) >= 3 and seg[2] == "rest"
        ]
    else:
        raw_pauses = [
            (datetime.fromisoformat(s), datetime.fromisoformat(e))
            for s, e in (act.pause_intervals or [])
        ]
    pieces = []
    cur = act.start_time
    while cur < act.end_time:
        next_midnight = datetime.combine(cur.date() + timedelta(days=1), datetime.min.time())
        piece_end = min(act.end_time, next_midnight)
        piece_pauses = []
        for p_start, p_end in raw_pauses:
            clipped_start, clipped_end = max(p_start, cur), min(p_end, piece_end)
            if clipped_start < clipped_end:
                piece_pauses.append([clipped_start.isoformat(), clipped_end.isoformat()])
        pieces.append(_DayPiece(
            start_time=cur,
            end_time=piece_end,
            pause_intervals=piece_pauses,
            activity_type=act.activity_type,
            salt_supplement=act.salt_supplement,
            vehicle_number=act.vehicle_number,
        ))
        cur = piece_end
    return pieces


# Medarbejdere med aftalen "timelønnet, ikke fastlagt arbejdstid" (agreement_kind
# hourly_flexible) har intet dagligt garanteret timetal – i stedet må de arbejde
# 37 "normal timer" i tidsrummet 06-18 PR. KALENDERUGE (mandag-søndag), hvorefter
# de næste 5 timer i samme tidsrum udløser OT 1-3-tillæg, og resten Øvrig overtid.
# Puljerne er ugentlige og deles derfor på tværs af alle dage/vagter i ugen, i
# modsætning til hourly_fixed, hvor loftet nulstilles hver dag (bekræftet af
# bruger 2026-08-17). Den ugentlige OT-1-3-pulje er fælles med aften-vinduets
# (18-21) tillæg, ligesom den daglige pulje er det for hourly_fixed.
WEEKLY_FLEX_NORMAL_MAX = Decimal("37")
WEEKLY_FLEX_OT13_MAX = Decimal("5")


def _calculate_employee(emp: Employee, start: date, end: date, db: Session) -> dict:
    """Beregn timefordeling og kr. for én medarbejder i et datointerval.
    Alle dage i perioden medtages – dage uden aktivitet vises som 0,
    fraværsdage vises med typenavn (beregning tilføjes senere)."""
    from collections import defaultdict
    from datetime import datetime as _dt

    activities = (
        db.query(Activity)
        .filter(
            Activity.employee_id == emp.id,
            Activity.status == ActivityStatus.approved,
            Activity.start_time < (end + timedelta(days=1)).isoformat(),
            Activity.end_time > start.isoformat(),
        )
        .order_by(Activity.start_time)
        .all()
    )

    try:
        hourly_rate = load_agreement_types_from_db(db).get(emp.agreement_type, Decimal("0"))
    except Exception as e:
        _logging.error(f"Timeløn kunne ikke indlæses for {emp.first_name} {emp.last_name} (id={emp.id}): {e}")
        raise HTTPException(500, f"Timeløn kunne ikke indlæses for {emp.first_name} {emp.last_name} – kontakt administrator")
    supplement = get_active_supplement_for_period(db, emp.id, start, end)
    if supplement:
        hourly_rate += supplement.value
    ot_rates = load_overtime_rates_from_db(db)
    salt_rate = load_salt_supplement_rate_from_db(db)
    overnight_rate = load_overnight_rate_from_db(db)
    dob_overnight_rate = load_dob_overnight_rate_from_db(db)
    dagpenge_sats = load_dagpenge_rate_from_db(db)
    springer_rate = load_springer_rate_from_db(db)
    ot_rates_by_id = load_overtime_rates_by_id_from_db(db)
    supplement_rates_by_id = load_supplement_rates_by_id_from_db(db)
    _springer_period = get_or_create_period_for_date(start, db)
    springer_enabled = db.query(EmployeeSpringerFlag).filter(
        EmployeeSpringerFlag.employee_id == emp.id,
        EmployeeSpringerFlag.pay_period_id == _springer_period.id,
        EmployeeSpringerFlag.enabled == True,
    ).first() is not None

    _ABSENCE_LABELS = {
        "ferie":        "Ferie",
        "fri":          "Fri",
        "afspadsering": "Afspadsering",
        "skole_kursus": "Skole/kursus",
        "sygdom":                  "Sygdom",
        "sygdom_u_8uger":          "Sygdom u. 8 uger",
        "sygdom_u_8_uger":         "Sygdom u. 8 uger",
        "barn_1sygedag":           "Barn 1.sygedag",
        "barn_1sygedag_u_8uger":   "Barn 1.sygedag u. 8 uger",
        "barn_2_3sygedag":         "Barn 2-3.sygedag",
        "paragraf_56_syg":         "§56 syg",
        "selvbetalt_fridag":       "Selvbetalt fridag",
        "feriefri":                "Feriefri",
        "barsel":                        "Barsel",
        "barsel_u_loen":                 "Barsel u. løn",
        "graviditetsbetinget_sygdom":    "Graviditetsbetinget sygdom",
    }

    # Indlæs helligdage for perioden (fra v14-helligdagskalender) – bruges også
    # til at afgøre om en aktivitet der strækker sig over midnat skal splittes.
    holiday_rows = db.query(Holiday).filter(
        Holiday.date >= start,
        Holiday.date <= end,
    ).all()
    holiday_map = {h.date: h for h in holiday_rows}

    _ABSOLUTE_DAY_TYPES = (
        DayType.SUNDAY, DayType.HOLIDAY_FULL,
        DayType.HOLIDAY_HALF_1MAJ, DayType.HOLIDAY_HALF_GRUNDLOV,
    )

    # Gruppér aktiviteter på dato (kan være flere pr. dag ved opdelinger).
    # Normaltids-/OT13-loftet hører til VAGTEN (den dag den startede) og
    # fortsætter uændret hen over midnat, så længe det ikke er brugt op – det
    # håndterer calculate_overtime() automatisk ved at behandle vagten som ét
    # sammenhængende opslag. Kun søndage/helligdage har en loft-uafhængig regel
    # ("alle kørte timer, uanset tidspunkt"), så en vagt der STARTER PÅ, eller
    # STRÆKKER SIG IND I (efter midnat), en søndag/helligdag SKAL splittes ved
    # midnat, så den del der falder på søndagen/helligdagen får dennes regel,
    # og resten falder tilbage til de øvrige dages egne (loft-baserede) regler.
    # Bekræftet af bruger 2026-07-02 (start-tilfældet) og 2026-08-17 (også når
    # vagten blot krydser ind i søndagen/helligdagen fra en tidligere hverdag/
    # lørdag – fx en lørdag-nattevagt der fortsætter forbi midnat ind i søndag).
    # Fraværstyper og overnatning er altid ét-dags og splittes aldrig.
    def _spans_absolute_day(act) -> bool:
        cur_d = act.start_time.date()
        end_d = act.end_time.date()
        while cur_d <= end_d:
            if classify_day(cur_d, holiday_map) in _ABSOLUTE_DAY_TYPES:
                return True
            cur_d += timedelta(days=1)
        return False

    acts_by_date = defaultdict(list)
    for act in activities:
        if act.activity_type in ("overnatning", "dob_overnatning") or _ABSENCE_LABELS.get(act.activity_type):
            acts_by_date[act.start_time.date()].append(act)
        elif _spans_absolute_day(act):
            for piece in _split_into_day_pieces(act):
                acts_by_date[piece.start_time.date()].append(piece)
        else:
            acts_by_date[act.start_time.date()].append(act)

    totals = {
        "normal": Decimal("0"), "ot_before": Decimal("0"),
        "ot_13": Decimal("0"), "ot_extra": Decimal("0"),
        "sh_kode8": Decimal("0"), "sh_kode9": Decimal("0"),
        "sh_fuldloennet": Decimal("0"), "sh_timeloennet": Decimal("0"),
        "afspadsering": Decimal("0"),
        "sygdom": Decimal("0"),
        "paragraf_56_syg": Decimal("0"),
        "barn_1sygedag_u_loen": Decimal("0"),
        "feriefri":     Decimal("0"),
        "barsel":       Decimal("0"),
        "skole_kursus": Decimal("0"),
        "salt_hours": Decimal("0"), "salt_kr": Decimal("0"),
        "overnight_count": 0,
        "dob_overnight_count": 0,
    }
    days = []
    total_kr = Decimal("0")

    # Overnatning håndteres som kolonne (ikke fraværsrække) – forhåndsberegn datoer.
    # DOB-overnatning tælles med i overnight_dates (samme dag-markering), men holdes
    # ude af overnight_count (kode 14) – den har sin egen tælling og sats (kode 43).
    overnight_dates = {a.start_time.date() for a in activities if a.activity_type in ("overnatning", "dob_overnatning")}
    totals["overnight_count"] = sum(1 for a in activities if a.activity_type == "overnatning")
    totals["dob_overnight_count"] = sum(1 for a in activities if a.activity_type == "dob_overnatning")

    is_recognized_agreement_kind = emp.agreement_kind in (
        AgreementKind.hourly_fixed, AgreementKind.hourly_flexible,
    )

    # hourly_flexible: 37t normaltid + 5t OT-1-3 er en ugentlig pulje (mandag-søndag),
    # ikke et dagligt loft – initialiseres her og videreføres/nulstilles i dag-løkken.
    is_hourly_flexible = emp.agreement_kind == AgreementKind.hourly_flexible
    week_normal_remaining = WEEKLY_FLEX_NORMAL_MAX
    week_ot13_remaining = WEEKLY_FLEX_OT13_MAX

    # hourly_fixed: når en vagt krydser midnat, nulstiller calculate_overtime nu
    # loftet internt til den nye kalenderdags eget loft (se next_day_normal_hours).
    # Den tilbageværende del af DET loft skal videreføres til næste dags bucket i
    # denne løkke, så en efterfølgende separat aktivitet SAMME (nye) kalenderdag
    # ikke fejlagtigt får sit eget friske loft oveni (dobbelt-tælling, fundet i
    # Anders Gervig Jensen-sagen 2026-08-17).
    midnight_carry = None

    # Gennemløb alle dage i perioden
    cur = start
    while cur <= end:
        acts_today = [a for a in acts_by_date.get(cur, []) if a.activity_type not in ("overnatning", "dob_overnatning")]
        overnight_today = 1 if cur in overnight_dates else 0

        # Dag-klassifikation og SH-betaling (gælder uanset om der køres,
        # MEDMINDRE medarbejderen er afløser og der ikke er kørsel den dag)
        day_type = classify_day(cur, holiday_map)
        guaranteed_today = _normal_hours_for_day(emp, cur)
        sh_h = compute_sh_hours(day_type, guaranteed_today)
        if emp.afloeser and not any(a.activity_type == "normal" for a in acts_today):
            sh_h = Decimal("0")
        if sh_h > 0:
            if emp.fuldloennet:
                totals["sh_fuldloennet"] += sh_h
            else:
                totals["sh_timeloennet"] += sh_h
            total_kr += sh_h * hourly_rate

        # Normaltids-/OT13-loft deles på tværs af dagens aktiviteter (ikke nulstillet
        # pr. aktivitet), så en dag der er delt i flere godkendte aktiviteter (fx efter
        # split) regnes som ét sammenhængende skift.
        if is_hourly_flexible:
            if cur.weekday() == 0:  # mandag – ny uge, ny 37t/5t-pulje
                week_normal_remaining = WEEKLY_FLEX_NORMAL_MAX
                week_ot13_remaining = WEEKLY_FLEX_OT13_MAX
            day_normal_remaining = week_normal_remaining
            day_ot13_remaining = week_ot13_remaining
        elif midnight_carry is not None and midnight_carry[0] == cur:
            # Gårsdagens vagt krydsede midnat ind i i dag – fortsæt med det loft,
            # den efterlod, i stedet for at give i dag et helt frisk loft oveni.
            _, day_normal_remaining, day_ot13_remaining = midnight_carry
            midnight_carry = None
        else:
            day_normal_remaining = guaranteed_today
            day_ot13_remaining = OT_13_MAX

        if not acts_today:
            days.append({
                "date": cur.isoformat(),
                "normal": 0.0, "ot_before": 0.0, "ot_13": 0.0,
                "ot_extra": 0.0, "total_hours": 0.0, "total_kr": 0.0,
                "absence_type": None,
                "start_time": None, "end_time": None,
                "vehicle_number": None,
                "overnight": overnight_today,
            })
        else:
            for act in acts_today:
                label = _ABSENCE_LABELS.get(act.activity_type)
                if label:
                    # absence_hours/absence_kr er KUN til visning i Lønafregning
                    # (payroll_settlement_router.py) – de indgår IKKE i totals{}
                    # eller total_kr her, og påvirker derfor ikke Lønkørsel,
                    # Excel-prøvekørslen, PDF-timesedlen eller Danløn-CSV'en.
                    absence_hours = None
                    absence_kr = None
                    if act.activity_type == "afspadsering":
                        dur = _afspadsering_hours(emp, act)
                        totals["afspadsering"] += dur
                        absence_hours = dur
                        absence_kr = dur * hourly_rate
                    elif act.activity_type == "ferie":
                        dur = Decimal(str((act.end_time - act.start_time).total_seconds())) / 3600
                        absence_hours = dur
                        absence_kr = dur * hourly_rate
                    elif act.activity_type in ("sygdom", "barn_1sygedag", "graviditetsbetinget_sygdom"):
                        dur = Decimal(str((act.end_time - act.start_time).total_seconds())) / 3600
                        totals["sygdom"] += dur
                        absence_hours = dur
                        absence_kr = dur * hourly_rate
                    elif act.activity_type == "paragraf_56_syg":
                        dur = Decimal(str((act.end_time - act.start_time).total_seconds())) / 3600
                        totals["paragraf_56_syg"] += dur
                        absence_hours = dur
                        absence_kr = dur * dagpenge_sats
                    elif act.activity_type == "feriefri":
                        dur = Decimal(str((act.end_time - act.start_time).total_seconds())) / 3600
                        totals["feriefri"] += dur
                        absence_hours = dur
                        # Fuldlønnet/timelønnet afgør kun hvilken Danløn-kode feriefri
                        # rapporteres under (FERIEFRI_FULDLOENNET/-TIMELOENNET, se
                        # export_csv_post) – selve beløbet regnes ens for begge
                        # (bekræftet af bruger 2026-08-25), så Lønafregning viser én
                        # samlet "Feriefri"-linje uanset emp.fuldloennet.
                        absence_kr = dur * hourly_rate
                    elif act.activity_type == "barsel":
                        dur = Decimal(str((act.end_time - act.start_time).total_seconds())) / 3600
                        totals["barsel"] += dur
                        absence_hours = dur
                        absence_kr = dur * hourly_rate
                    elif act.activity_type == "barn_1sygedag_u_8uger":
                        dur = Decimal(str((act.end_time - act.start_time).total_seconds())) / 3600
                        totals["barn_1sygedag_u_loen"] += dur
                        absence_hours = dur
                        absence_kr = dur * dagpenge_sats
                    elif act.activity_type == "skole_kursus":
                        dur = Decimal(str((act.end_time - act.start_time).total_seconds())) / 3600
                        totals["skole_kursus"] += dur
                        absence_hours = dur
                        absence_kr = dur * hourly_rate
                        # Skole/kursus-timer forbruger dagens garanterede timer ligesom
                        # arbejdstid, så efterfølgende kørsel samme dag korrekt bliver
                        # overtid når garantien er brugt op (bekræftet af bruger for
                        # Mikkel Bo Rosenkilde, mandag 24/8-2026).
                        day_normal_remaining = max(Decimal("0"), day_normal_remaining - dur)
                    elif act.activity_type == "sygdom_u_8uger":
                        # Ulønnet – vises med timer i Lønafregning, men altid 0 kr.
                        absence_hours = Decimal(str((act.end_time - act.start_time).total_seconds())) / 3600
                        absence_kr = Decimal("0")
                    # barn_2_3sygedag / selvbetalt_fridag / barsel_u_loen: ikke i CSV
                    days.append({
                        "date": cur.isoformat(),
                        "normal": 0.0, "ot_before": 0.0, "ot_13": 0.0,
                        "ot_extra": 0.0, "total_hours": 0.0, "total_kr": 0.0,
                        "absence_type": label,
                        "absence_hours": float(absence_hours) if absence_hours is not None else None,
                        "absence_kr": float(_round2(absence_kr)) if absence_kr is not None else None,
                        "start_time": None, "end_time": None,
                        "vehicle_number": None,
                        "overnight": overnight_today,
                    })
                else:
                    _segs = getattr(act, "segments", None)
                    if _segs:
                        pauses = [
                            (_dt.fromisoformat(seg[0]), _dt.fromisoformat(seg[1]))
                            for seg in _segs if len(seg) >= 3 and seg[2] == "rest"
                        ]
                    else:
                        pauses = [
                            (_dt.fromisoformat(s), _dt.fromisoformat(e))
                            for s, e in (act.pause_intervals or [])
                        ]
                    pause_minutes = sum(
                        (Decimal(str((p_end - p_start).total_seconds())) / Decimal("60")
                         for p_start, p_end in pauses),
                        Decimal("0"),
                    )
                    if not is_recognized_agreement_kind:
                        # Aftale-type uden for de to kendte nøgler – ingen
                        # automatisk OT-beregning endnu (se
                        # docs/superpowers/specs/2026-08-24-aftale-stamdata-design.md).
                        ot = calculate_flat_hours(act.start_time, act.end_time, pauses)
                    elif day_type in (DayType.NORMAL, DayType.SATURDAY):
                        # Lørdag er ikke længere en særlig dag – den bruger samme
                        # tidsvindues-beregning som en hverdag, med lørdagens egne
                        # garanterede timer (typisk 0) som loft (bekræftet 2026-07-02).
                        # hourly_fixed har et DAGLIGT loft, så en vagt der krydser
                        # midnat skal skifte til den nye kalenderdags eget loft (og
                        # friske 3t OT-1-3) – ellers "arver" fx lørdagens andel af en
                        # fredagsvagt fejlagtigt fredagens loft (bekræftet 2026-08-17).
                        # hourly_flexible's ugentlige pulje må IKKE nulstilles ved
                        # midnat, så next_day_normal_hours udelades for den.
                        _next_day_normal = (
                            None if is_hourly_flexible
                            else _normal_hours_for_day(emp, cur + timedelta(days=1))
                        )
                        ot = calculate_overtime(
                            act.start_time, act.end_time,
                            guaranteed_today, pauses, ot_rates,
                            normal_remaining=day_normal_remaining,
                            ot13_remaining=day_ot13_remaining,
                            next_day_normal_hours=_next_day_normal,
                        )
                        day_normal_remaining = ot.normal_remaining_after
                        day_ot13_remaining = ot.ot13_remaining_after
                        if not is_hourly_flexible and act.end_time.date() != cur:
                            midnight_carry = (act.end_time.date(), day_normal_remaining, day_ot13_remaining)
                    else:
                        ot = calculate_special_day_overtime(
                            act.start_time, act.end_time,
                            day_type, pauses,
                            kode8_remaining=day_ot13_remaining,
                        )
                        day_ot13_remaining = ot.ot13_remaining_after
                    day_salt_hours = ot.total_hours if act.salt_supplement else Decimal("0")
                    day_salt_kr = day_salt_hours * salt_rate
                    day_kr = (
                        ot.total_hours * hourly_rate
                        + ot.ot_before_hours * ot_rates[OT_BEFORE_KEY]
                        + ot.ot_13_hours * ot_rates[OT_13_KEY]
                        + ot.ot_extra_hours * ot_rates[OT_EXTRA_KEY]
                        + ot.sh_kode8_hours * ot_rates[OT_13_KEY]
                        + ot.sh_kode9_hours * ot_rates[OT_EXTRA_KEY]
                        + day_salt_kr
                    )
                    totals["normal"]     += ot.normal_hours
                    totals["ot_before"]  += ot.ot_before_hours
                    totals["ot_13"]      += ot.ot_13_hours
                    totals["ot_extra"]   += ot.ot_extra_hours
                    totals["sh_kode8"]   += ot.sh_kode8_hours
                    totals["sh_kode9"]   += ot.sh_kode9_hours
                    totals["salt_hours"] += day_salt_hours
                    totals["salt_kr"]    += day_salt_kr
                    total_kr             += day_kr
                    days.append({
                        "date": cur.isoformat(),
                        "normal":       float(_round2(ot.normal_hours)),
                        "ot_before":    float(_round2(ot.ot_before_hours)),
                        "ot_13":        float(_round2(ot.ot_13_hours)),
                        "ot_extra":     float(_round2(ot.ot_extra_hours)),
                        "total_hours":  float(_round2(ot.total_hours)),
                        "total_kr":     float(_round2(day_kr)),
                        "pause_minutes": float(pause_minutes.quantize(Decimal("1"), rounding=ROUND_HALF_UP)),
                        "absence_type": None,
                        "sh_kode8":     float(_round2(ot.sh_kode8_hours)),
                        "sh_kode9":     float(_round2(ot.sh_kode9_hours)),
                        "salt_hours":   float(_round2(day_salt_hours)),
                        "salt_kr":      float(_round2(day_salt_kr)),
                        "start_time":   act.start_time.strftime("%H:%M"),
                        "end_time":     act.end_time.strftime("%H:%M"),
                        "vehicle_number": act.vehicle_number or "",
                        "overnight":    overnight_today,
                    })
        if is_hourly_flexible:
            week_normal_remaining = day_normal_remaining
            week_ot13_remaining = day_ot13_remaining
        cur += timedelta(days=1)

    has_pending = (
        db.query(Activity)
        .filter(
            Activity.employee_id == emp.id,
            Activity.status == ActivityStatus.pending,
            Activity.start_time >= start.isoformat(),
            Activity.start_time < (end + timedelta(days=1)).isoformat(),
        )
        .count() > 0
    )

    # normal_hours indeholder nu alle arbejdede timer (tillæg er additive)
    total_hours = totals["normal"]

    return {
        "employee_id":        emp.id,
        "employee_number":    emp.employee_number,
        "employee_name":      emp.name,
        "email":              emp.email,
        "agreement_type":     emp.agreement_type,
        "hourly_rate":        float(hourly_rate),
        "salt_rate":          float(salt_rate),
        "ot_rates":           {k: float(v) for k, v in ot_rates.items()},
        "ot_rates_by_id":     {k: float(v) for k, v in ot_rates_by_id.items()},
        "supplement_rates_by_id": {k: float(v) for k, v in supplement_rates_by_id.items()},
        "days":               days,
        "normal_hours":       float(_round2(totals["normal"])),
        "ot_before_hours":    float(_round2(totals["ot_before"])),
        "ot_13_hours":        float(_round2(totals["ot_13"])),
        "ot_extra_hours":     float(_round2(totals["ot_extra"])),
        "salt_hours":         float(_round2(totals["salt_hours"])),
        "salt_kr":            float(_round2(totals["salt_kr"])),
        "overnight_count":    totals["overnight_count"],
        "overnight_rate":     float(overnight_rate),
        "overnight_kr":       float(_round2(Decimal(str(totals["overnight_count"])) * overnight_rate)),
        "dob_overnight_count": totals["dob_overnight_count"],
        "dob_overnight_rate":  float(dob_overnight_rate),
        "dob_overnight_kr":    float(_round2(Decimal(str(totals["dob_overnight_count"])) * dob_overnight_rate)),
        "springer_rate":      float(springer_rate),
        "springer_enabled":   springer_enabled,
        "afspadsering_hours":   float(_round2(totals["afspadsering"])),
        "sygdom_hours":         float(_round2(totals["sygdom"])),
        "paragraf_56_syg_hours":  float(_round2(totals["paragraf_56_syg"])),
        "barn_1sygedag_u_loen_hours": float(_round2(totals["barn_1sygedag_u_loen"])),
        "feriefri_hours":         float(_round2(totals["feriefri"])),
        "barsel_hours":           float(_round2(totals["barsel"])),
        "skole_kursus_hours":     float(_round2(totals["skole_kursus"])),
        "dagpenge_sats":          float(dagpenge_sats),
        "total_hours":        float(_round2(total_hours)),
        "total_kr":                float(_round2(total_kr)),
        "activity_count":          len(activities),
        "has_pending":             has_pending,
        "sh_fuldloennet_hours":    float(_round2(totals["sh_fuldloennet"])),
        "sh_timeloennet_hours":    float(_round2(totals["sh_timeloennet"])),
        "sh_kode8_hours":          float(_round2(totals["sh_kode8"])),
        "sh_kode9_hours":          float(_round2(totals["sh_kode9"])),
    }


def _resolve_period(period_start: Optional[str], db: Session):
    d = date.fromisoformat(period_start) if period_start else date.today()
    return get_or_create_period_for_date(d, db)


def _active_employees(db: Session, employee_id: Optional[int] = None):
    q = db.query(Employee).filter(Employee.active == True)
    if employee_id:
        q = q.filter(Employee.id == employee_id)
    employees = q.order_by(Employee.first_name, Employee.last_name).all()
    return [e for e in employees if e.dispatcher_group and e.dispatcher_group.visible_in_activity_overview]


@router.get("/preview")
def payroll_preview(period_start: Optional[str] = None,
                    current_user: AppUser = Depends(_payroll_access),
                    db: Session = Depends(get_db)):
    period = _resolve_period(period_start, db)
    employees = _active_employees(db)
    results = [_calculate_employee(e, period.start_date, period.end_date, db) for e in employees]
    return {
        "period_start": period.start_date.isoformat(),
        "period_end": period.end_date.isoformat(),
        "period_status": period.status.value,
        "employees": results,
        "has_unresolved_pending": any(r["has_pending"] for r in results),
    }


def _build_proevekoersel_workbook(employees, period, db):
    """Bygger prøvekørsel-Excel-arbejdsbog og returnerer den (fælles for download og gem)."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prøvekørsel"
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="317423", end_color="317423", fill_type="solid")
    row_fill = PatternFill(start_color="D4EDCC", end_color="D4EDCC", fill_type="solid")

    headers = ["Medarbejder", "Lønnr", "Vognnr.", "Dag", "Starttid", "Sluttid", "Total tid",
               "Pause i alt (min)", "Normal tid", "Overtid 1 time før", "Overtid 1-3 timer efter",
               "Øvrig overtid", "Salttillæg (t)", "Salttillæg (kr.)", "Overnatning", "Total kr."]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        # Overskrifterne er længere end de fleste dataceller (fx "Overtid
        # 1-3 timer efter") – ombryd dem i stedet for at gøre kolonnen
        # unødvendigt bred, så hele tabellen kan ses uden vandret scroll.
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for emp in employees:
        calc = _calculate_employee(emp, period.start_date, period.end_date, db)
        # Markør så brugeren kan se med det samme, at medarbejderen har fået
        # springertillæg i denne periode, uden at det kræver en ekstra kolonne.
        emp_name = calc["employee_name"] + (" (springer)" if calc.get("springer_enabled") else "")
        first_row = True
        for day in calc["days"]:
            vn = day.get("vehicle_number") or ""
            st = day.get("start_time") or ""
            et = day.get("end_time") or ""
            on = day.get("overnight", 0) or ""
            if day["absence_type"]:
                row = [emp_name, calc["employee_number"], "", day["date"],
                       "", "", "", "", day["absence_type"], "", "", "", "", "",
                       on, ""]
            else:
                salt_t = day.get("salt_hours", 0) or ""
                salt_kr = day.get("salt_kr", 0) or ""
                day_ot13 = day["ot_13"] + day.get("sh_kode8", 0)
                day_ot_extra = day["ot_extra"] + day.get("sh_kode9", 0)
                # "Normal tid" skal kun være timer der IKKE udløser overtidstillæg.
                # day["normal"] indeholder alle arbejdede timer (tillæggene er
                # additive oveni, se calculators/overtime.py), så tillægstimerne
                # trækkes fra her for at vise den rene normaltid.
                day_normal = round(day["normal"] - day["ot_before"] - day_ot13 - day_ot_extra, 2) + 0.0
                row = [emp_name, calc["employee_number"], vn, day["date"],
                       st, et,
                       day["total_hours"], day.get("pause_minutes", 0.0),
                       day_normal, day["ot_before"],
                       day_ot13,
                       day_ot_extra,
                       salt_t, salt_kr, on,
                       day["total_kr"]]
            ws.append(row)
            if first_row:
                for cell in ws[ws.max_row]:
                    cell.fill = row_fill
                first_row = False
        total_ot13 = calc["ot_13_hours"] + calc.get("sh_kode8_hours", 0)
        total_ot_extra = calc["ot_extra_hours"] + calc.get("sh_kode9_hours", 0)
        total_normal = round(calc["normal_hours"] - calc["ot_before_hours"] - total_ot13 - total_ot_extra, 2) + 0.0
        total_pause = sum(d.get("pause_minutes", 0.0) for d in calc["days"])
        total_row = [emp_name, calc["employee_number"], "", "TOTAL",
                     "", "",
                     calc["total_hours"], total_pause, total_normal, calc["ot_before_hours"],
                     total_ot13,
                     total_ot_extra,
                     calc.get("salt_hours", 0) or "", calc.get("salt_kr", 0) or "",
                     calc.get("overnight_count", 0) or "",
                     calc["total_kr"]]
        ws.append(total_row)
        for cell in ws[ws.max_row]:
            cell.font = bold
        sh_fl = calc.get("sh_fuldloennet_hours", 0)
        sh_tl = calc.get("sh_timeloennet_hours", 0)
        hr = calc["hourly_rate"]
        if sh_fl > 0:
            sh_row = [emp_name, calc["employee_number"], "", "Søgnehelligdag",
                      "", "", sh_fl, "", sh_fl, "", "", "", "", "", "", round(sh_fl * hr, 2)]
            ws.append(sh_row)
            for cell in ws[ws.max_row]:
                cell.font = bold
        if sh_tl > 0:
            sh_row = [emp_name, calc["employee_number"], "", "SH-Udbetaling",
                      "", "", sh_tl, "", sh_tl, "", "", "", "", "", "", round(sh_tl * hr, 2)]
            ws.append(sh_row)
            for cell in ws[ws.max_row]:
                cell.font = bold
        on_kr = calc.get("overnight_kr", 0.0)
        dagpenge = calc.get("dagpenge_sats", 137.43)
        for abs_lbl, abs_h, abs_rate in [
            ("Sygdom med løn",  calc.get("sygdom_hours", 0),              hr),
            ("§56 syg",         calc.get("paragraf_56_syg_hours", 0),     dagpenge),
            ("Barn 1.sygedag",  calc.get("barn_1sygedag_u_loen_hours", 0), dagpenge),
            ("Feriefri",        calc.get("feriefri_hours", 0),            hr),
            ("Barsel",          calc.get("barsel_hours", 0),              hr),
            ("Kursus/Skole",    calc.get("skole_kursus_hours", 0),        hr),
        ]:
            if abs_h > 0:
                ws.append([emp_name, calc["employee_number"], "", abs_lbl,
                           "", "", abs_h, "", abs_h, "", "", "", "", "", "", round(abs_h * abs_rate, 2)])
                for cell in ws[ws.max_row]:
                    cell.font = bold
        if on_kr > 0:
            ws.append([emp_name, calc["employee_number"], "", "Overnatning (kr.)",
                       "", "", "", "", "", "", "", "", "", "", "", round(on_kr, 2)])
            for cell in ws[ws.max_row]:
                cell.font = bold
        dob_on_kr = calc.get("dob_overnight_kr", 0.0)
        if dob_on_kr > 0:
            ws.append([emp_name, calc["employee_number"], "", "DOB Overnatning (kr.)",
                       "", "", "", "", "", "", "", "", "", "", "", round(dob_on_kr, 2)])
            for cell in ws[ws.max_row]:
                cell.font = bold
        ws.append([])

    # Bredder tilpasset dataindholdet (ikke de – ofte længere – ombrudte
    # overskrifter), så alle 16 kolonner kan ses uden vandret scroll.
    column_widths = {
        "A": 24,  # Medarbejder
        "B": 8,   # Lønnr
        "C": 9,   # Vognnr.
        "D": 12,  # Dag
        "E": 8,   # Starttid
        "F": 8,   # Sluttid
        "G": 9,   # Total tid
        "H": 10,  # Pause i alt (min)
        "I": 9,   # Normal tid
        "J": 9,   # Overtid 1 time før
        "K": 10,  # Overtid 1-3 timer efter
        "L": 9,   # Øvrig overtid
        "M": 9,   # Salttillæg (t)
        "N": 10,  # Salttillæg (kr.)
        "O": 10,  # Overnatning
        "P": 10,  # Total kr.
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    return wb


@router.get("/proevekoersel")
def proevekoersel(
    period_start: Optional[str] = None,
    employee_id: Optional[int] = None,
    current_user: AppUser = Depends(_payroll_access),
    db: Session = Depends(get_db),
):
    """Prøvekørsel: Excel-fil med mellemregninger – alle eller én medarbejder."""
    period = _resolve_period(period_start, db)
    employees = _active_employees(db, employee_id)
    if not employees:
        raise HTTPException(404, "Ingen medarbejdere fundet")

    wb = _build_proevekoersel_workbook(employees, period, db)
    filename = f"proevekoersel_{period.start_date.isoformat()}_{period.end_date.isoformat()}.xlsx"
    OUTPUT_DIR.mkdir(exist_ok=True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    (OUTPUT_DIR / filename).write_bytes(buf.getvalue())
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


class ProevekoerselSaveRequest(BaseModel):
    period_start: Optional[str] = None
    employee_id: Optional[int] = None
    output_folder: str


@router.post("/proevekoersel-gem")
def proevekoersel_gem(body: ProevekoerselSaveRequest,
                      current_user: AppUser = Depends(_payroll_access),
                      db: Session = Depends(get_db)):
    """Prøvekørsel gemt til valgt mappe i stedet for browser-download."""
    period = _resolve_period(body.period_start, db)
    employees = _active_employees(db, body.employee_id)
    if not employees:
        raise HTTPException(404, "Ingen medarbejdere fundet")

    wb = _build_proevekoersel_workbook(employees, period, db)
    save_dir = _safe_save_dir(body.output_folder)
    try:
        save_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        _logging.error(f"Kan ikke oprette mappe '{save_dir}': {exc}")
        raise HTTPException(400, "Mappen kunne ikke oprettes – tjek stien og rettigheder")

    filename = f"proevekoersel_{period.start_date.isoformat()}_{period.end_date.isoformat()}.xlsx"
    filepath = save_dir / filename
    try:
        wb.save(str(filepath))
    except PermissionError:
        raise HTTPException(
            400,
            f"Kunne ikke gemme filen '{filename}' – tjek om den er åben i Excel eller et andet program, og prøv igen.",
        )
    return {"path": str(filepath), "filename": filename}


@router.get("/export-csv")
def export_csv(period_start: Optional[str] = None,
               current_user: AppUser = Depends(_payroll_access),
               db: Session = Depends(get_db)):
    """
    Kør løn: Danløn CSV.
    Kolonner: A=CVR, B=medarbejdernummer, C=Danløn-kode,
              D=antal timer, E=time-/tillægssats.
    """
    period = _resolve_period(period_start, db)
    employees = _active_employees(db)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", lineterminator="\r\n")

    def fmt(v: float) -> str:
        return str(round(v * 100))

    pt = _get_pay_type_data(db)
    _code     = lambda k: pt.get(k, {}).get("code", "1")
    _in_csv   = lambda k: pt.get(k, {}).get("in_csv", True)
    _inc_rate = lambda k: pt.get(k, {}).get("inc_rate", True)
    _inc_tot  = lambda k: pt.get(k, {}).get("inc_total", False)

    for emp in employees:
        calc = _calculate_employee(emp, period.start_date, period.end_date, db)
        if (calc["activity_count"] == 0
                and calc["afspadsering_hours"] == 0
                and calc["sygdom_hours"] == 0
                and calc["paragraf_56_syg_hours"] == 0
                and calc["barn_1sygedag_u_loen_hours"] == 0
                and calc["feriefri_hours"] == 0
                and calc["barsel_hours"] == 0
                and calc["skole_kursus_hours"] == 0
                and calc.get("sh_fuldloennet_hours", 0) == 0
                and calc.get("sh_timeloennet_hours", 0) == 0):
            continue

        raw_rows = [
            ("NORMAL",         calc["normal_hours"],                                               calc["hourly_rate"]),
            _springer_row(calc),
            ("OT_BEFORE",      calc["ot_before_hours"],                                            calc["ot_rates"][OT_BEFORE_KEY]),
            ("OT_13",          calc["ot_13_hours"] + calc.get("sh_kode8_hours", 0),               calc["ot_rates"][OT_13_KEY]),
            ("OT_EXTRA",       calc["ot_extra_hours"] + calc.get("sh_kode9_hours", 0),            calc["ot_rates"][OT_EXTRA_KEY]),
            ("SH_FULDLOENNET", calc.get("sh_fuldloennet_hours", 0),                               calc["hourly_rate"]),
            ("SH_TIMELOENNET", calc.get("sh_timeloennet_hours", 0),                               calc["hourly_rate"]),
            ("SALT",           calc.get("salt_hours", 0),                                         calc.get("salt_rate", 0)),
            ("OVERNATNING",    calc.get("overnight_count", 0),                                    calc.get("overnight_rate", 0)),
            ("AFSPADSERING",   calc["afspadsering_hours"],                                        calc["hourly_rate"]),
            ("SYGDOM",         calc["sygdom_hours"],                                              calc["hourly_rate"]),
            ("PARAGRAF_56",    calc["paragraf_56_syg_hours"],                                     calc.get("dagpenge_sats", 137.43)),
            ("BARN_1SYGEDAG",  calc["barn_1sygedag_u_loen_hours"],                                calc.get("dagpenge_sats", 137.43)),
            ("FERIEFRI",       _builtin_absence_qty(pt, "FERIEFRI", "feriefri", calc["feriefri_hours"],
                                                      emp.id, period.start_date, period.end_date, db), calc["hourly_rate"]),
            *([("FERIEFRI_FULDLOENNET",  calc["feriefri_hours"], calc["hourly_rate"])] if emp.fuldloennet else []),
            *([("FERIEFRI_TIMELOENNET", calc["feriefri_hours"], calc["hourly_rate"])] if not emp.fuldloennet else []),
            ("BARSEL",         calc["barsel_hours"],                                              calc["hourly_rate"]),
            ("SKOLE_KURSUS",   calc["skole_kursus_hours"],                                        calc["hourly_rate"]),
        ] + _user_pay_type_rows(emp.id, period.start_date, period.end_date, calc, db)
        code_agg = {}
        for key, qty, rate in raw_rows:
            if not _in_csv(key) or qty == 0:
                continue
            code = _code(key)
            if code in code_agg:
                prev_qty, prev_rate, prev_inc_rate, prev_inc_tot = code_agg[code]
                code_agg[code] = (prev_qty + qty, prev_rate, prev_inc_rate, prev_inc_tot)
            else:
                code_agg[code] = (qty, rate, _inc_rate(key), _inc_tot(key))
        for code, (qty, rate, inc_rate, inc_tot) in code_agg.items():
            qty_fmt = fmt(qty)
            row = [_get_employee_cvr(emp, db), calc["employee_number"], code, qty_fmt]
            row.append(fmt(rate) if inc_rate else "")
            row.append(fmt(qty * float(rate)) if inc_tot else "")
            writer.writerow(row)

    filename = f"danloen_{period.start_date.isoformat()}_{period.end_date.isoformat()}.csv"
    OUTPUT_DIR.mkdir(exist_ok=True)
    (OUTPUT_DIR / filename).write_bytes(output.getvalue().encode("utf-8"))

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


class ExportCsvRequest(BaseModel):
    period_start: Optional[str] = None
    output_folder: str


@router.post("/export-csv")
def export_csv_post(body: ExportCsvRequest,
                    current_user: AppUser = Depends(_payroll_access),
                    db: Session = Depends(get_db)):
    """
    Kør løn: låser perioden og gemmer Danløn CSV til valgt mappe.
    """
    period = _resolve_period(body.period_start, db)

    if period.status == PayPeriodStatus.closed:
        raise HTTPException(
            400,
            "Lønperioden er allerede låst – lønnen er allerede kørt for denne periode. "
            "Åbn perioden igen under Administration, hvis der skal foretages ændringer.",
        )

    pending_count = (
        db.query(Activity)
        .join(Employee, Activity.employee_id == Employee.id)
        .filter(
            Employee.active == True,
            Activity.status == ActivityStatus.pending,
            Activity.start_time >= period.start_date.isoformat(),
            Activity.start_time < (period.end_date + timedelta(days=1)).isoformat(),
        )
        .count()
    )
    if pending_count > 0:
        raise HTTPException(
            400,
            "Lønnen kan ikke køres – der er aktiviteter, der afventer godkendelse eller deaktivering.",
        )

    employees = _active_employees(db)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", lineterminator="\r\n")

    def fmt(v: float) -> str:
        return str(round(v * 100))

    pt = _get_pay_type_data(db)
    _code     = lambda k: pt.get(k, {}).get("code", "1")
    _in_csv   = lambda k: pt.get(k, {}).get("in_csv", True)
    _inc_rate = lambda k: pt.get(k, {}).get("inc_rate", True)
    _inc_tot  = lambda k: pt.get(k, {}).get("inc_total", False)

    for emp in employees:
        calc = _calculate_employee(emp, period.start_date, period.end_date, db)
        if (calc["activity_count"] == 0
                and calc["afspadsering_hours"] == 0
                and calc["sygdom_hours"] == 0
                and calc["paragraf_56_syg_hours"] == 0
                and calc["barn_1sygedag_u_loen_hours"] == 0
                and calc["feriefri_hours"] == 0
                and calc["barsel_hours"] == 0
                and calc["skole_kursus_hours"] == 0
                and calc.get("sh_fuldloennet_hours", 0) == 0
                and calc.get("sh_timeloennet_hours", 0) == 0):
            continue
        raw_rows = [
            ("NORMAL",         calc["normal_hours"],                                               calc["hourly_rate"]),
            _springer_row(calc),
            ("OT_BEFORE",      calc["ot_before_hours"],                                            calc["ot_rates"][OT_BEFORE_KEY]),
            ("OT_13",          calc["ot_13_hours"] + calc.get("sh_kode8_hours", 0),               calc["ot_rates"][OT_13_KEY]),
            ("OT_EXTRA",       calc["ot_extra_hours"] + calc.get("sh_kode9_hours", 0),            calc["ot_rates"][OT_EXTRA_KEY]),
            ("SH_FULDLOENNET", calc.get("sh_fuldloennet_hours", 0),                               calc["hourly_rate"]),
            ("SH_TIMELOENNET", calc.get("sh_timeloennet_hours", 0),                               calc["hourly_rate"]),
            ("SALT",           calc.get("salt_hours", 0),                                         calc.get("salt_rate", 0)),
            ("OVERNATNING",    calc.get("overnight_count", 0),                                    calc.get("overnight_rate", 0)),
            ("AFSPADSERING",   calc["afspadsering_hours"],                                        calc["hourly_rate"]),
            ("SYGDOM",         calc["sygdom_hours"],                                              calc["hourly_rate"]),
            ("PARAGRAF_56",    calc["paragraf_56_syg_hours"],                                     calc.get("dagpenge_sats", 137.43)),
            ("BARN_1SYGEDAG",  calc["barn_1sygedag_u_loen_hours"],                                calc.get("dagpenge_sats", 137.43)),
            ("FERIEFRI",       _builtin_absence_qty(pt, "FERIEFRI", "feriefri", calc["feriefri_hours"],
                                                      emp.id, period.start_date, period.end_date, db), calc["hourly_rate"]),
            *([("FERIEFRI_FULDLOENNET",  calc["feriefri_hours"], calc["hourly_rate"])] if emp.fuldloennet else []),
            *([("FERIEFRI_TIMELOENNET", calc["feriefri_hours"], calc["hourly_rate"])] if not emp.fuldloennet else []),
            ("BARSEL",         calc["barsel_hours"],                                              calc["hourly_rate"]),
            ("SKOLE_KURSUS",   calc["skole_kursus_hours"],                                        calc["hourly_rate"]),
        ] + _user_pay_type_rows(emp.id, period.start_date, period.end_date, calc, db)
        code_agg = {}
        for key, qty, rate in raw_rows:
            if not _in_csv(key) or qty == 0:
                continue
            code = _code(key)
            if code in code_agg:
                prev_qty, prev_rate, prev_inc_rate, prev_inc_tot = code_agg[code]
                code_agg[code] = (prev_qty + qty, prev_rate, prev_inc_rate, prev_inc_tot)
            else:
                code_agg[code] = (qty, rate, _inc_rate(key), _inc_tot(key))
        for code, (qty, rate, inc_rate, inc_tot) in code_agg.items():
            qty_fmt = fmt(qty)
            row = [_get_employee_cvr(emp, db), calc["employee_number"], code, qty_fmt]
            row.append(fmt(rate) if inc_rate else "")
            row.append(fmt(qty * float(rate)) if inc_tot else "")
            writer.writerow(row)

    filename = f"danloen_{period.start_date.isoformat()}_{period.end_date.isoformat()}.csv"
    save_dir = _safe_save_dir(body.output_folder)
    try:
        save_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        import logging; logging.error(f"Kan ikke oprette mappe '{save_dir}': {exc}")
        raise HTTPException(400, "Mappen kunne ikke oprettes – tjek stien og rettigheder")
    try:
        (save_dir / filename).write_bytes(output.getvalue().encode("utf-8"))
    except PermissionError:
        raise HTTPException(
            400,
            f"Kunne ikke gemme filen '{filename}' – tjek om den er åben i Excel eller et andet program, og prøv igen.",
        )

    period.status = PayPeriodStatus.closed
    period.closed_at = datetime.utcnow()
    period.closed_by = current_user.initials
    log_action(db, current_user, "payroll_run", "pay_period", period.id,
               f"Løn kørt for periode {period.start_date} – {period.end_date}")
    db.commit()

    return {"filename": filename, "path": str(save_dir / filename)}


@router.post("/reopen-period")
def reopen_period(period_start: Optional[str] = None,
                  current_user: AppUser = Depends(_reopen_access),
                  db: Session = Depends(get_db)):
    """Åbner en lukket lønperiode igen (sætter status tilbage til 'open')."""
    period = _resolve_period(period_start, db)
    period.status = PayPeriodStatus.open
    period.closed_at = None
    period.closed_by = None

    # Aktiviteter der blev oprettet manuelt mens perioden var lukket, ruller automatisk
    # frem til næste åbne periode (se get_billing_period i calculators/pay_period.py).
    # Nu hvor perioden genåbnes, skal de aktiviteter der reelt hører til denne periodes
    # datointerval, men fejlagtigt peger på en senere periode, flyttes tilbage - ellers
    # bliver de usynlige i Aktiviteter-fanen for begge perioder.
    start_str = period.start_date.isoformat()
    end_str = (period.end_date + timedelta(days=1)).isoformat()
    stray = (
        db.query(Activity)
        .filter(
            Activity.pay_period_id != period.id,
            Activity.start_time >= start_str,
            Activity.start_time < end_str,
        )
        .all()
    )
    for a in stray:
        a.pay_period_id = period.id
    if stray:
        log_action(db, current_user, "reassign_pay_period_id", "pay_period", period.id,
                   f"{len(stray)} aktivitet(er) flyttet tilbage til periode {period.start_date} – {period.end_date} ved genåbning")

    log_action(db, current_user, "reopen_period", "pay_period", period.id,
               f"Lønperiode genåbnet: {period.start_date} – {period.end_date}")
    db.commit()
    return {"status": "open", "period_start": period.start_date.isoformat()}


class PdfRequest(BaseModel):
    from_date: date
    to_date: date
    employee_id: Optional[int] = None
    output_folder: Optional[str] = None


@router.get("/downloads-folder")
def get_downloads_folder(current_user: AppUser = Depends(_payroll_access)):
    """Returnerer brugerens Downloads-mappe som forslag til gem-placering."""
    from pathlib import Path as _P
    folder = _P.home() / "Downloads"
    return {"path": str(folder)}


@router.post("/pdf-timesedler")
def pdf_timesedler(body: PdfRequest,
                   current_user: AppUser = Depends(_payroll_access),
                   db: Session = Depends(get_db)):
    """
    Dan PDF-timesedler for valgt datointerval.
    PDF'erne gemmes i output/timesedler/ (e-mail-afsendelse tilføjes senere).
    Bruger samme layout som /api/timeseddel (routers/timeseddel_router.py _build_pdf).
    """
    from routers.timeseddel_router import _build_pdf

    if body.to_date < body.from_date:
        raise HTTPException(400, "Til-dato skal være efter fra-dato")

    employees = _active_employees(db, body.employee_id)
    if body.output_folder:
        pdf_dir = _safe_save_dir(body.output_folder)
    else:
        pdf_dir = OUTPUT_DIR / "timesedler"
    try:
        pdf_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        import logging; logging.error(f"Kan ikke oprette mappe '{pdf_dir}': {exc}")
        raise HTTPException(400, "Mappen kunne ikke oprettes – tjek stien og rettigheder")

    created = []
    skipped = []

    for emp in employees:
        calc = _calculate_employee(emp, body.from_date, body.to_date, db)
        if calc["activity_count"] == 0:
            skipped.append(calc["employee_name"])
            continue

        filename = f"timeseddel_{emp.employee_number}_{body.from_date.isoformat()}_{body.to_date.isoformat()}.pdf"
        path = pdf_dir / filename

        pdf_bytes = _build_pdf(calc, _get_employee_cvr(emp, db))
        path.write_bytes(pdf_bytes)
        created.append({"employee": calc["employee_name"], "email": calc["email"], "file": str(path)})

    return {
        "created": created,
        "skipped": skipped,
        "folder": str(pdf_dir),
        "note": "PDF'er gemt lokalt – e-mail-afsendelse er ikke konfigureret endnu.",
    }
