"""
Indlæser satser fra Excel-arkene i projektets rodmappe:
- "Overtid satser.xlsx": de tre overtidstillæg
- "Overenskomsttyper og timesatser.xlsx": overenskomsttyper med timesatser

Arkene vedligeholdes af lønafdelingen og genindlæses ved hvert kald,
så ændringer slår igennem uden genstart.
"""
import shutil
import tempfile
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Optional

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
OVERTIME_RATES_FILE = BASE_DIR / "Overtid satser.xlsx"
AGREEMENT_TYPES_FILE = BASE_DIR / "Overenskomsttyper og timesatser.xlsx"
SALT_SUPPLEMENT_FILE = BASE_DIR / "Salttillæg og overnatning.xlsx"


def _load_workbook_safe(path: Path):
    """OneDrive/Excel kan låse filen – læs via en midlertidig kopi."""
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except PermissionError:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)
        shutil.copy2(path, tmp_path)
        try:
            return openpyxl.load_workbook(tmp_path, data_only=True)
        finally:
            tmp_path.unlink(missing_ok=True)


def load_overtime_rates() -> dict[str, Decimal]:
    """
    Returnerer {'Overtid 1 time før': Decimal, 'Overtid 1-3 timer efter': ...,
    'Øvrigt overtid': ...} fra "Overtid satser.xlsx".
    """
    if not OVERTIME_RATES_FILE.exists():
        raise FileNotFoundError(f"Satsefil mangler: {OVERTIME_RATES_FILE}")
    wb = _load_workbook_safe(OVERTIME_RATES_FILE)
    ws = wb.worksheets[0]
    rates = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        rates[str(row[0]).strip()] = Decimal(str(row[1]))
    return rates


def load_agreement_types() -> dict[str, Decimal]:
    """
    Returnerer {overenskomsttype: timesats} fra
    "Overenskomsttyper og timesatser.xlsx" (kolonne A og B).
    """
    if not AGREEMENT_TYPES_FILE.exists():
        raise FileNotFoundError(f"Satsefil mangler: {AGREEMENT_TYPES_FILE}")
    wb = _load_workbook_safe(AGREEMENT_TYPES_FILE)
    ws = wb.worksheets[0]
    types = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        types[str(row[0]).strip()] = Decimal(str(row[1]))
    return types


def load_salt_supplement_rate() -> Decimal:
    """
    Returnerer salttillæg pr. time fra "Salttillæg og overnatning.xlsx".
    Søger i kolonne A efter rækken med teksten "salttillæg" og returnerer
    værdien i kolonne B — rækkefølgen i arket er ligegyldig.
    """
    if not SALT_SUPPLEMENT_FILE.exists():
        raise FileNotFoundError(f"Satsefil mangler: {SALT_SUPPLEMENT_FILE}")
    wb = _load_workbook_safe(SALT_SUPPLEMENT_FILE)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(values_only=True):
        label = str(row[0]).strip().lower() if row[0] is not None else ""
        if label == "salttillæg":
            if row[1] is None:
                raise ValueError("Salttillæg-satsen mangler i kolonne B")
            return Decimal(str(row[1]))
    raise ValueError('Rækken "Salttillæg" blev ikke fundet i Salttillæg og overnatning.xlsx')


def load_overnight_rate() -> Decimal:
    """
    Returnerer overnatning-sats fra "Salttillæg og overnatning.xlsx".
    Søger i kolonne A efter rækken med teksten "overnatning" og returnerer
    værdien i kolonne B — rækkefølgen i arket er ligegyldig.
    """
    if not SALT_SUPPLEMENT_FILE.exists():
        raise FileNotFoundError(f"Satsefil mangler: {SALT_SUPPLEMENT_FILE}")
    wb = _load_workbook_safe(SALT_SUPPLEMENT_FILE)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(values_only=True):
        label = str(row[0]).strip().lower() if row[0] is not None else ""
        if label == "overnatning":
            if row[1] is None:
                raise ValueError("Overnatning-satsen mangler i kolonne B")
            return Decimal(str(row[1]))
    raise ValueError('Rækken "Overnatning" blev ikke fundet i Salttillæg og overnatning.xlsx')


def get_hourly_rate(agreement_type: str) -> Decimal:
    types = load_agreement_types()
    if agreement_type not in types:
        raise KeyError(f"Ukendt overenskomsttype: {agreement_type}")
    return types[agreement_type]


def seniority_variant_exists(agreement_type: str) -> str | None:
    """
    Returnerer navnet på 9-mdr-anciennitetsvarianten af en overenskomsttype,
    hvis den findes i arket – ellers None.
    """
    types = load_agreement_types()
    candidate = f"{agreement_type}. 9 mdr anciennitet"
    return candidate if candidate in types else None


# ── DB-baserede versioner (bruger stamdata-tabeller som primær kilde) ──────


def load_agreement_types_from_db(db) -> dict[str, Decimal]:
    from database.models import MasterAgreementType
    rows = db.query(MasterAgreementType).all()
    if rows:
        return {r.name: Decimal(str(r.hourly_rate)) for r in rows}
    return load_agreement_types()


def load_overtime_rates_from_db(db) -> dict[str, Decimal]:
    from database.models import MasterOvertimeRate
    rows = db.query(MasterOvertimeRate).all()
    if rows:
        return {r.label: Decimal(str(r.rate)) for r in rows}
    return load_overtime_rates()


def load_salt_supplement_rate_from_db(db) -> Decimal:
    from database.models import MasterSupplementRate
    row = db.query(MasterSupplementRate).filter(MasterSupplementRate.label == "Salttillæg").first()
    if row:
        return Decimal(str(row.rate))
    try:
        return load_salt_supplement_rate()
    except Exception:
        return Decimal("0")


def load_overnight_rate_from_db(db) -> Decimal:
    from database.models import MasterSupplementRate
    row = db.query(MasterSupplementRate).filter(MasterSupplementRate.label == "Overnatning").first()
    if row:
        return Decimal(str(row.rate))
    try:
        return load_overnight_rate()
    except Exception:
        return Decimal("0")


def load_springer_rate_from_db(db) -> Decimal:
    from database.models import MasterSupplementRate
    row = db.query(MasterSupplementRate).filter(MasterSupplementRate.label == "Springertillæg").first()
    return Decimal(str(row.rate)) if row else Decimal("0")


def load_dagpenge_rate_from_db(db) -> Decimal:
    from database.models import MasterSupplementRate
    from calculators.pay_rates import DANLOEN_DAGPENGE_SATS
    row = db.query(MasterSupplementRate).filter(MasterSupplementRate.label == "Dagpenge §56").first()
    if row:
        return Decimal(str(row.rate))
    return DANLOEN_DAGPENGE_SATS


def load_dob_overnight_rate_from_db(db) -> Decimal:
    """DOB-overnatningens tillægssats – brugeroprettet via Stamdata → Tillæg,
    intet Excel-fallback (i modsætning til load_overnight_rate_from_db)."""
    from database.models import MasterSupplementRate
    row = db.query(MasterSupplementRate).filter(MasterSupplementRate.label == "DOB_overnatning").first()
    return Decimal(str(row.rate)) if row else Decimal("0")


def load_overtime_rates_by_id_from_db(db) -> dict[int, Decimal]:
    """Alle overtidssatser nøglet på id (ikke kun de tre faste) — bruges af
    _resolve_rate() til at slå brugerdefinerede sats-kilder op dynamisk."""
    from database.models import MasterOvertimeRate
    return {r.id: Decimal(str(r.rate)) for r in db.query(MasterOvertimeRate).all()}


def load_supplement_rates_by_id_from_db(db) -> dict[int, Decimal]:
    """Alle tillægssatser nøglet på id (ikke kun de fast navngivne) — bruges af
    _resolve_rate() til at slå brugerdefinerede sats-kilder op dynamisk."""
    from database.models import MasterSupplementRate
    return {r.id: Decimal(str(r.rate)) for r in db.query(MasterSupplementRate).all()}


def seniority_variant_exists_from_db(db, agreement_type: str) -> str | None:
    types = load_agreement_types_from_db(db)
    candidate = f"{agreement_type}. 9 mdr anciennitet"
    return candidate if candidate in types else None


def get_active_supplement_for_period(
    db, employee_id: int, period_start: date, period_end: date
) -> Optional["EmployeeSupplement"]:
    """Finder tillægget hvis gyldighedsperiode overlapper [period_start, period_end].
    Overlapper flere rækker (nyt tillæg oprettet midt i perioden), vinder den
    med nyeste start_date, for hele perioden."""
    from database.models import EmployeeSupplement
    return (
        db.query(EmployeeSupplement)
        .filter(
            EmployeeSupplement.employee_id == employee_id,
            EmployeeSupplement.end_date >= period_start,
            EmployeeSupplement.start_date <= period_end,
        )
        .order_by(EmployeeSupplement.start_date.desc())
        .first()
    )
