import logging
from pathlib import Path

from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker

BASE_DIR = Path(__file__).resolve().parent.parent
DB_PATH = BASE_DIR / "database" / "lonsystem.db"

engine = create_engine(
    f"sqlite:///{DB_PATH}",
    connect_args={"check_same_thread": False},
)


@event.listens_for(engine, "connect")
def set_wal_mode(dbapi_connection, connection_record):
    try:
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA journal_mode=WAL")
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()
    except Exception as e:
        logging.error(f"WAL-mode opsætning fejlede: {e}")


SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


def get_db():
    db = SessionLocal()
    try:
        yield db
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def init_db():
    from database.models import Base
    Base.metadata.create_all(bind=engine)
    _migrate()
    _seed_roles()
    _seed_admin()
    _seed_master_data()
    _seed_cvr()
    _seed_holidays()
    _ensure_sh_pay_types()    # SH-løntypekoder kode 4 og 63
    _ensure_anciennitet_alert_permission()
    _ensure_paragraf_56_alert_permission()
    _ensure_manage_baselines_permission()
    _ensure_activity_permissions()
    _ensure_auto_approve_permission()
    _ensure_manage_auto_approval_permission()
    _ensure_system_settings()
    _ensure_vagtplan_permissions()
    _ensure_employee_supplements_permission()
    _ensure_toggle_springer_permission()
    _ensure_payroll_settlement_permissions()
    _ensure_springer_pay_type()
    _ensure_feriefri_fuldloennet_pay_type()
    _migrate_dispatcher_groups()
    _migrate_dispatcher_group_to_single()


def _migrate():
    """Tilføjer nye kolonner til eksisterende databaser uden at miste data."""
    import sqlite3 as _sqlite3
    with _sqlite3.connect(str(DB_PATH)) as conn:
        existing = {row[1] for row in conn.execute("PRAGMA table_info(activities)")}
        if "created_by" not in existing:
            conn.execute("ALTER TABLE activities ADD COLUMN created_by VARCHAR")
            conn.commit()
        ot_cols = {row[1] for row in conn.execute("PRAGMA table_info(master_overtime_rates)")}
        if "is_user_created" not in ot_cols:
            conn.execute("ALTER TABLE master_overtime_rates ADD COLUMN is_user_created BOOLEAN DEFAULT 0")
            conn.commit()
        sup_cols = {row[1] for row in conn.execute("PRAGMA table_info(master_supplement_rates)")}
        if "is_user_created" not in sup_cols:
            conn.execute("ALTER TABLE master_supplement_rates ADD COLUMN is_user_created BOOLEAN DEFAULT 0")
            conn.commit()
        pt_cols = {row[1] for row in conn.execute("PRAGMA table_info(master_pay_types)")}
        if "is_user_created" not in pt_cols:
            conn.execute("ALTER TABLE master_pay_types ADD COLUMN is_user_created BOOLEAN DEFAULT 0")
            conn.commit()
        if "csv_quantity_type" not in pt_cols:
            conn.execute("ALTER TABLE master_pay_types ADD COLUMN csv_quantity_type VARCHAR(20) NOT NULL DEFAULT 'hours'")
            conn.execute("ALTER TABLE master_pay_types ADD COLUMN csv_rate_source VARCHAR(30) NOT NULL DEFAULT 'hourly'")
            conn.execute("ALTER TABLE master_pay_types ADD COLUMN csv_include_rate BOOLEAN NOT NULL DEFAULT 1")
            conn.execute("ALTER TABLE master_pay_types ADD COLUMN csv_include_total BOOLEAN NOT NULL DEFAULT 0")
            conn.commit()
            conn.execute("UPDATE master_pay_types SET csv_rate_source='ot_before' WHERE code_key='OT_BEFORE'")
            conn.execute("UPDATE master_pay_types SET csv_rate_source='ot_13' WHERE code_key='OT_13'")
            conn.execute("UPDATE master_pay_types SET csv_rate_source='ot_extra' WHERE code_key='OT_EXTRA'")
            conn.execute("UPDATE master_pay_types SET csv_rate_source='salt' WHERE code_key='SALT'")
            conn.execute("UPDATE master_pay_types SET csv_quantity_type='count', csv_rate_source='overnight' WHERE code_key='OVERNATNING'")
            conn.execute("UPDATE master_pay_types SET csv_rate_source='dagpenge' WHERE code_key='PARAGRAF_56'")
            conn.execute("UPDATE master_pay_types SET csv_rate_source='dagpenge' WHERE code_key='BARN_1SYGEDAG'")
            conn.commit()
        if "csv_include_rate" not in pt_cols:
            conn.execute("ALTER TABLE master_pay_types ADD COLUMN csv_include_rate BOOLEAN NOT NULL DEFAULT 1")
            conn.commit()
        emp_cols = {row[1] for row in conn.execute("PRAGMA table_info(employees)")}
        if "cvr_number" not in emp_cols:
            conn.execute("ALTER TABLE employees ADD COLUMN cvr_number VARCHAR(20)")
            conn.commit()
        if "anciennitet_dismissed_at" not in emp_cols:
            conn.execute("ALTER TABLE employees ADD COLUMN anciennitet_dismissed_at DATETIME")
            conn.commit()
        if "terminsdato" not in emp_cols:
            conn.execute("ALTER TABLE employees ADD COLUMN terminsdato DATE")
            conn.commit()
        if "initials" not in emp_cols:
            conn.execute("ALTER TABLE employees ADD COLUMN initials VARCHAR(10)")
            conn.commit()
        if "dispatcher_group_id" not in emp_cols:
            conn.execute("ALTER TABLE employees ADD COLUMN dispatcher_group_id INTEGER")
            conn.commit()
        if "paragraf_56" not in emp_cols:
            conn.execute("ALTER TABLE employees ADD COLUMN paragraf_56 BOOLEAN NOT NULL DEFAULT 0")
            conn.execute("ALTER TABLE employees ADD COLUMN paragraf_56_start_date DATE")
            conn.execute("ALTER TABLE employees ADD COLUMN paragraf_56_end_date DATE")
            conn.commit()
        if "afloeser" not in emp_cols:
            conn.execute("ALTER TABLE employees ADD COLUMN afloeser BOOLEAN NOT NULL DEFAULT 0")
            conn.commit()
        act_cols2 = {row[1] for row in conn.execute("PRAGMA table_info(activities)")}
        if "deactivated_by" not in act_cols2:
            conn.execute("ALTER TABLE activities ADD COLUMN deactivated_by VARCHAR")
            conn.commit()
        if "auto_approved" not in act_cols2:
            conn.execute("ALTER TABLE activities ADD COLUMN auto_approved BOOLEAN NOT NULL DEFAULT 0")
            conn.commit()
        if "auto_approval_flags" not in act_cols2:
            conn.execute("ALTER TABLE activities ADD COLUMN auto_approval_flags TEXT NOT NULL DEFAULT '[]'")
            conn.commit()
        if "is_likely_incomplete" not in act_cols2:
            conn.execute("ALTER TABLE activities ADD COLUMN is_likely_incomplete BOOLEAN NOT NULL DEFAULT 0")
            conn.commit()
        if "baseline_duration_minutes" not in act_cols2:
            conn.execute("ALTER TABLE activities ADD COLUMN baseline_duration_minutes REAL")
            conn.commit()
        if "baseline_start_hour" not in act_cols2:
            conn.execute("ALTER TABLE activities ADD COLUMN baseline_start_hour REAL")
            conn.commit()
        if "hidden_from_vagtplan" not in act_cols2:
            conn.execute("ALTER TABLE activities ADD COLUMN hidden_from_vagtplan BOOLEAN NOT NULL DEFAULT 0")
            conn.commit()
        existing_indexes = {row[1] for row in conn.execute("PRAGMA index_list(activities)")}
        if "ix_activities_employee_start_source" not in existing_indexes:
            conn.execute(
                "CREATE INDEX ix_activities_employee_start_source "
                "ON activities(employee_id, start_time, source)"
            )
            conn.commit()
        if "uq_activities_employee_start_tachograph" not in existing_indexes:
            # Fjern evt. eksisterende dubletter (samme medarbejder+starttid,
            # tachograf-kilde) opstået pga. den tidligere manglende spærre,
            # før det unikke indeks oprettes – ellers fejler CREATE UNIQUE
            # INDEX på en database, hvor racet allerede er indtruffet.
            # Beholder den ældste (laveste id); den nyeste vinder normalt
            # ikke noget ekstra data, da genimport allerede opdaterer den
            # eksisterende række (se import_ddd.py::_import_activity).
            conn.execute(
                "DELETE FROM activities WHERE source = 'tachograph' AND id NOT IN ("
                "  SELECT MIN(id) FROM activities WHERE source = 'tachograph' "
                "  GROUP BY employee_id, start_time"
                ")"
            )
            conn.execute(
                "CREATE UNIQUE INDEX uq_activities_employee_start_tachograph "
                "ON activities(employee_id, start_time) WHERE source = 'tachograph'"
            )
            conn.commit()
        if "ix_activities_period_status" not in existing_indexes:
            conn.execute(
                "CREATE INDEX ix_activities_period_status ON activities(pay_period_id, status)"
            )
            conn.commit()
        baseline_indexes = {row[1] for row in conn.execute("PRAGMA index_list(employee_baselines)")}
        if "uq_employee_baselines_employee_weekday" not in baseline_indexes:
            # Fjern evt. eksisterende dubletter (samme medarbejder+ugedag,
            # opstået pga. den tidligere manglende spærre) før det unikke
            # indeks oprettes. Beholder rækken med flest samples – den har
            # det mest velfunderede datagrundlag af de to.
            conn.execute(
                "DELETE FROM employee_baselines WHERE id NOT IN ("
                "  SELECT id FROM ("
                "    SELECT id, ROW_NUMBER() OVER ("
                "      PARTITION BY employee_id, weekday ORDER BY sample_count DESC, id ASC"
                "    ) AS rn FROM employee_baselines"
                "  ) WHERE rn = 1"
                ")"
            )
            conn.execute(
                "CREATE UNIQUE INDEX uq_employee_baselines_employee_weekday "
                "ON employee_baselines(employee_id, weekday)"
            )
            conn.commit()
        audit_indexes = {row[1] for row in conn.execute("PRAGMA index_list(audit_logs)")}
        if "ix_audit_logs_timestamp" not in audit_indexes:
            conn.execute("CREATE INDEX ix_audit_logs_timestamp ON audit_logs(timestamp)")
            conn.commit()
        dg_cols = {row[1] for row in conn.execute("PRAGMA table_info(dispatcher_groups)")}
        if "vehicle_id" not in dg_cols:
            conn.execute("ALTER TABLE dispatcher_groups ADD COLUMN vehicle_id INTEGER")
            conn.commit()
        if "visible_in_activity_overview" not in dg_cols:
            conn.execute(
                "ALTER TABLE dispatcher_groups ADD COLUMN visible_in_activity_overview "
                "BOOLEAN NOT NULL DEFAULT 1"
            )
            conn.commit()
        conn.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS uq_employee_supplements_one_open_row "
            "ON employee_supplements(employee_id) WHERE end_date = '9999-12-31'"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS ix_employee_supplements_employee_id "
            "ON employee_supplements(employee_id)"
        )
        conn.commit()

        # Migrer eksisterende faste sats-kilde-værdier til det nye id-baserede skema
        # (overtime:<id> / supplement:<id>) – idempotent, rammer kun rækker der
        # stadig har en af de gamle faste værdier.
        _legacy_rate_src_map = {
            "ot_before": ("master_overtime_rates", "Overtid 1 time før"),
            "ot_13": ("master_overtime_rates", "Overtid 1-3 timer efter"),
            "ot_extra": ("master_overtime_rates", "Øvrigt overtid"),
            "salt": ("master_supplement_rates", "Salttillæg"),
            "overnight": ("master_supplement_rates", "Overnatning"),
            "dagpenge": ("master_supplement_rates", "Dagpenge §56"),
            "springer": ("master_supplement_rates", "Springertillæg"),
        }
        for old_value, (table, label) in _legacy_rate_src_map.items():
            found = conn.execute(f"SELECT id FROM {table} WHERE label = ?", (label,)).fetchone()
            if found:
                prefix = "overtime" if table == "master_overtime_rates" else "supplement"
                conn.execute(
                    "UPDATE master_pay_types SET csv_rate_source = ? WHERE csv_rate_source = ?",
                    (f"{prefix}:{found[0]}", old_value),
                )
        conn.commit()


def _seed_roles():
    from database.models import Role
    db = SessionLocal()
    try:
        if db.query(Role).count() == 0:
            for r in [
                Role(name="admin", display_name="Administrator", is_system=True,
                     permissions=["payroll", "import_ddd", "user_management", "reopen_period", "manage_baselines", "manage_auto_approval", "approve_activities", "view_calendar"]),
                Role(name="lonbogholder", display_name="Lønbogholder", is_system=False,
                     permissions=["payroll", "absence_overview", "import_ddd", "anciennitet_alert", "approve_activities", "view_calendar", "auto_approve_manual_activities"]),
                Role(name="disponent", display_name="Disponent", is_system=False,
                     permissions=["approve_activities", "view_calendar"]),
            ]:
                db.add(r)
            db.commit()
    finally:
        db.close()


def _seed_admin():
    from database.models import AppUser
    from auth import hash_password
    db = SessionLocal()
    try:
        if db.query(AppUser).count() == 0:
            db.add(AppUser(
                name="Administrator",
                initials="admin",
                email="",
                role="admin",
                password_hash=hash_password("admin"),
                active=True,
            ))
            db.commit()
    finally:
        db.close()


def _normalize_absence_key(label: str) -> str:
    overrides = {"Kursus/Skole": "skole_kursus"}
    if label in overrides:
        return overrides[label]
    s = label.lower()
    s = s.replace("§", "paragraf_")
    s = s.replace("æ", "ae").replace("ø", "oe").replace("å", "aa")
    s = s.replace(" ", "_").replace("/", "_").replace(".", "").replace("-", "_")
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")


def _seed_agreement_kinds(db):
    from database.models import MasterAgreementKind
    if db.query(MasterAgreementKind).count() == 0:
        db.add(MasterAgreementKind(
            key="hourly_fixed", label="Timelønnet, fast arbejdstid",
            is_active=True, is_user_created=False,
            requires_agreement_type=True, sort_order=1,
        ))
        db.add(MasterAgreementKind(
            key="hourly_flexible", label="Timelønnet, ikke fastlagt arbejdstid",
            is_active=True, is_user_created=False,
            requires_agreement_type=True, sort_order=2,
        ))
        db.commit()


def _seed_master_data():
    from decimal import Decimal
    from database.models import (
        MasterAgreementType, MasterOvertimeRate,
        MasterSupplementRate, MasterPayType, MasterAbsenceType,
        MasterAgreementKind,
    )
    db = SessionLocal()
    try:
        if db.query(MasterAgreementType).count() == 0:
            try:
                from calculators.rates_loader import load_agreement_types
                for name, rate in load_agreement_types().items():
                    db.add(MasterAgreementType(name=name, hourly_rate=rate))
                db.commit()
            except Exception as e:
                db.rollback()
                logging.warning(f"Overenskomsttyper – seeding fra Excel slog fejl: {e}")

        if db.query(MasterOvertimeRate).count() == 0:
            try:
                from calculators.rates_loader import load_overtime_rates
                for label, rate in load_overtime_rates().items():
                    db.add(MasterOvertimeRate(label=label, rate=rate))
                db.commit()
            except Exception as e:
                db.rollback()
                logging.warning(f"Overtidssatser – seeding fra Excel slog fejl: {e}")

        if db.query(MasterSupplementRate).count() == 0:
            try:
                from calculators.rates_loader import load_salt_supplement_rate, load_overnight_rate
                db.add(MasterSupplementRate(label="Salttillæg", rate=load_salt_supplement_rate()))
                db.add(MasterSupplementRate(label="Overnatning", rate=load_overnight_rate()))
                db.add(MasterSupplementRate(label="Dagpenge §56", rate=Decimal("137.43")))
                db.commit()
            except Exception as e:
                db.rollback()
                logging.warning(f"Tillægssatser – seeding slog fejl: {e}")

        if db.query(MasterPayType).count() == 0:
            from calculators.pay_rates import (
                DANLOEN_CODE_NORMAL, DANLOEN_CODE_OT_BEFORE, DANLOEN_CODE_OT_13,
                DANLOEN_CODE_OT_EXTRA, DANLOEN_CODE_SALT, DANLOEN_CODE_AFSPADSERING,
                DANLOEN_CODE_SYGDOM, DANLOEN_CODE_FERIEFRI, DANLOEN_CODE_BARSEL,
                DANLOEN_CODE_SKOLE_KURSUS, DANLOEN_CODE_OVERNATNING,
                DANLOEN_CODE_PARAGRAF_56, DANLOEN_CODE_BARN_1SYGEDAG,
            )
            pay_types = [
                # (code_key, label, danloen_code, include_in_csv, sort_order, qty_type, rate_src)
                ("NORMAL",        "Normal tid",           DANLOEN_CODE_NORMAL,        True,  1, "hours", "hourly"),
                ("OT_BEFORE",     "Overtid 1 time før",   DANLOEN_CODE_OT_BEFORE,     True,  2, "hours", "ot_before"),
                ("OT_13",         "Overtid 1-3 timer",    DANLOEN_CODE_OT_13,         True,  3, "hours", "ot_13"),
                ("OT_EXTRA",      "Øvrig overtid",        DANLOEN_CODE_OT_EXTRA,      True,  4, "hours", "ot_extra"),
                ("SALT",          "Salttillæg",           DANLOEN_CODE_SALT,          True,  5, "hours", "salt"),
                ("OVERNATNING",   "Overnatning",          DANLOEN_CODE_OVERNATNING,   True,  6, "count", "overnight"),
                ("AFSPADSERING",  "Afspadsering",         DANLOEN_CODE_AFSPADSERING,  True,  7, "hours", "hourly"),
                ("SYGDOM",        "Sygdom med løn",       DANLOEN_CODE_SYGDOM,        True,  8, "hours", "hourly"),
                ("PARAGRAF_56",   "§56 syg",              DANLOEN_CODE_PARAGRAF_56,   True,  9, "hours", "dagpenge"),
                ("BARN_1SYGEDAG", "Barn 1.sygedag",       DANLOEN_CODE_BARN_1SYGEDAG, True, 10, "hours", "dagpenge"),
                ("FERIEFRI",      "Feriefri",             DANLOEN_CODE_FERIEFRI,      True, 11, "hours", "hourly"),
                ("BARSEL",        "Barsel",               DANLOEN_CODE_BARSEL,        True, 12, "hours", "hourly"),
                ("SKOLE_KURSUS",  "Kursus/Skole",         DANLOEN_CODE_SKOLE_KURSUS,  True, 13, "hours", "hourly"),
            ]
            for ck, lbl, code, inc, order, qty, rate_src in pay_types:
                db.add(MasterPayType(
                    code_key=ck, label=lbl, danloen_code=code,
                    include_in_csv=inc, sort_order=order,
                    csv_quantity_type=qty, csv_rate_source=rate_src,
                    csv_include_rate=True, csv_include_total=False,
                ))
            db.commit()
        if db.query(MasterAbsenceType).count() == 0:
            try:
                from openpyxl import load_workbook
                xlsx_path = BASE_DIR / "Fraværstyper.xlsx"
                _backend_only = {
                    "sygdom_u_8uger", "sygdom_u_8_uger",
                    "barn_1sygedag_u_8uger", "barsel_u_loen",
                }
                if xlsx_path.exists():
                    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
                    ws = wb.active
                    order = 1
                    first = True
                    for row in ws.iter_rows(values_only=True):
                        if first:
                            first = False
                            continue
                        cell = row[0]
                        if cell:
                            lbl = str(cell).strip()
                            key = _normalize_absence_key(lbl)
                            if key not in _backend_only:
                                db.add(MasterAbsenceType(
                                    label=lbl, normalized_key=key,
                                    is_active=True, is_user_created=False,
                                    sort_order=order,
                                ))
                                order += 1
                    wb.close()
                    db.commit()
            except Exception as e:
                db.rollback()
                logging.warning(f"Fraværstyper – seeding fra Excel slog fejl: {e}")

        _seed_agreement_kinds(db)

    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved seeding af stamdata: {e}")
    finally:
        db.close()


def _seed_cvr():
    from database.models import MasterCvrNumber
    from calculators.pay_rates import CVR_NUMBER
    db = SessionLocal()
    try:
        if db.query(MasterCvrNumber).count() == 0:
            db.add(MasterCvrNumber(
                cvr_number=CVR_NUMBER,
                company_name="Poul Schou A/S",
                is_default=True,
            ))
            db.commit()
    finally:
        db.close()


def _seed_holidays():
    from database.models import Holiday
    from calculators.holidays import get_holidays_for_year
    from datetime import date
    db = SessionLocal()
    try:
        current_year = date.today().year
        for year in range(current_year, current_year + 5):
            for h in get_holidays_for_year(year):
                if not db.query(Holiday).filter(Holiday.date == h["date"]).first():
                    db.add(Holiday(
                        date=h["date"],
                        name=h["name"],
                        half_day_from=h["half_day_from"],
                        is_auto_generated=True,
                    ))
        db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved seeding af helligdage: {e}")
    finally:
        db.close()


def _ensure_sh_pay_types():
    """Tilføjer SH-løntypekoder til eksisterende databaser (idempotent)."""
    from database.models import MasterPayType
    from calculators.pay_rates import DANLOEN_CODE_SH_FULDLOENNET, DANLOEN_CODE_SH_TIMELOENNET
    db = SessionLocal()
    try:
        entries = [
            ("SH_FULDLOENNET", "Søgnehelligdag", DANLOEN_CODE_SH_FULDLOENNET, True, 14, "hours", "hourly"),
            ("SH_TIMELOENNET", "SH-Udbetaling", DANLOEN_CODE_SH_TIMELOENNET, True, 15, "hours", "hourly"),
        ]
        for ck, lbl, code, inc, order, qty, rate_src in entries:
            existing = db.query(MasterPayType).filter(MasterPayType.code_key == ck).first()
            if not existing:
                db.add(MasterPayType(
                    code_key=ck, label=lbl, danloen_code=code,
                    include_in_csv=inc, sort_order=order,
                    csv_quantity_type=qty, csv_rate_source=rate_src,
                    csv_include_rate=True, csv_include_total=False,
                ))
            elif existing.label != lbl:
                existing.label = lbl
        db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved seeding af SH-løntypekoder: {e}")
    finally:
        db.close()


def _ensure_springer_pay_type():
    """Seeder Springertillæg-sats og -løntypekode til eksisterende databaser (idempotent)."""
    from decimal import Decimal
    from database.models import MasterSupplementRate, MasterPayType
    from calculators.pay_rates import DANLOEN_CODE_SPRINGERTILLAEG
    db = SessionLocal()
    try:
        if not db.query(MasterSupplementRate).filter(MasterSupplementRate.label == "Springertillæg").first():
            db.add(MasterSupplementRate(label="Springertillæg", rate=Decimal("0")))
        if not db.query(MasterPayType).filter(MasterPayType.code_key == "SPRINGERTILLAEG").first():
            db.add(MasterPayType(
                code_key="SPRINGERTILLAEG", label="Springertillæg",
                danloen_code=DANLOEN_CODE_SPRINGERTILLAEG,
                include_in_csv=True, sort_order=16,
                csv_quantity_type="hours", csv_rate_source="springer",
                csv_include_rate=True, csv_include_total=False,
            ))
        db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved seeding af Springertillæg: {e}")
    finally:
        db.close()


def _ensure_feriefri_fuldloennet_pay_type():
    """Seeder FERIEFRI_FULDLOENNET-løntypekode til eksisterende databaser (idempotent)."""
    from database.models import MasterPayType
    from calculators.pay_rates import DANLOEN_CODE_FERIEFRI_FULDLOENNET
    db = SessionLocal()
    try:
        if not db.query(MasterPayType).filter(MasterPayType.code_key == "FERIEFRI_FULDLOENNET").first():
            db.add(MasterPayType(
                code_key="FERIEFRI_FULDLOENNET", label="Feriefri fuldlønnet",
                danloen_code=DANLOEN_CODE_FERIEFRI_FULDLOENNET,
                include_in_csv=True, sort_order=17,
                csv_quantity_type="hours", csv_rate_source="hourly",
                csv_include_rate=True, csv_include_total=False,
            ))
        db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved seeding af FERIEFRI_FULDLOENNET-løntypekode: {e}")
    finally:
        db.close()


def _ensure_anciennitet_alert_permission():
    """Tilføjer anciennitet_alert til lonbogholder-rollen (idempotent)."""
    from database.models import Role
    db = SessionLocal()
    try:
        role = db.query(Role).filter(Role.name == "lonbogholder").first()
        if role and not role.is_system:
            perms = list(role.permissions or [])
            if "anciennitet_alert" not in perms:
                perms.append("anciennitet_alert")
                role.permissions = perms
                db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved opdatering af anciennitet_alert-tilladelse: {e}")
    finally:
        db.close()


def _ensure_paragraf_56_alert_permission():
    """Tilføjer paragraf_56_alert til lonbogholder-rollen (idempotent)."""
    from database.models import Role
    db = SessionLocal()
    try:
        role = db.query(Role).filter(Role.name == "lonbogholder").first()
        if role and not role.is_system:
            perms = list(role.permissions or [])
            if "paragraf_56_alert" not in perms:
                perms.append("paragraf_56_alert")
                role.permissions = perms
                db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved opdatering af paragraf_56_alert-tilladelse: {e}")
    finally:
        db.close()


def _migrate_dispatcher_groups():
    """
    Seeder de faste disponentgrupper og migrerer eksisterende medarbejderes
    (legacy) dispatcher_group-streng til den nye many-to-many-tabel.
    Idempotent – dropper legacy-kolonnen efter migrering.
    """
    import sqlite3 as _sqlite3
    from database.models import DispatcherGroup

    db = SessionLocal()
    try:
        default_groups = [
            "2 - Kran", "4 - Makulering", "5 - Miljø",
            "8 - THG", "9 - BN", "10 - ISOPLUS-CHJ",
        ]
        existing_names = {g.name for g in db.query(DispatcherGroup).all()}
        for name in default_groups:
            if name not in existing_names:
                db.add(DispatcherGroup(name=name))
        db.commit()

        with _sqlite3.connect(str(DB_PATH)) as conn:
            emp_cols = {row[1] for row in conn.execute("PRAGMA table_info(employees)")}
            if "dispatcher_group" not in emp_cols:
                return
            rows = conn.execute(
                "SELECT id, dispatcher_group FROM employees "
                "WHERE dispatcher_group IS NOT NULL AND dispatcher_group != ''"
            ).fetchall()
            groups_by_name = {g.name: g.id for g in db.query(DispatcherGroup).all()}
            for emp_id, group_name in rows:
                group_id = groups_by_name.get(group_name)
                if group_id is None:
                    g = DispatcherGroup(name=group_name)
                    db.add(g)
                    db.commit()
                    db.refresh(g)
                    groups_by_name[group_name] = g.id
                    group_id = g.id
                already_linked = conn.execute(
                    "SELECT 1 FROM employee_dispatcher_groups WHERE employee_id = ? AND dispatcher_group_id = ?",
                    (emp_id, group_id),
                ).fetchone()
                if not already_linked:
                    conn.execute(
                        "INSERT INTO employee_dispatcher_groups (employee_id, dispatcher_group_id) VALUES (?, ?)",
                        (emp_id, group_id),
                    )
            conn.commit()
            conn.execute("ALTER TABLE employees DROP COLUMN dispatcher_group")
            conn.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved migrering af disponentgrupper: {e}")
    finally:
        db.close()


def _migrate_dispatcher_group_to_single():
    """
    Reducerer disponentgruppe fra mange-til-mange til én gruppe pr. medarbejder.
    Har en medarbejder i dag flere grupper, beholdes den alfabetisk først
    sorterede (efter gruppenavn). Idempotent – dropper employee_dispatcher_groups
    efter migrering; kører derfor kun data-trinnet én gang (tabellen er væk bagefter).
    Kolonnerne dispatcher_group_id/vehicle_id oprettes af _migrate(), som kører
    før denne funktion, så de findes allerede når vi når hertil.
    """
    import sqlite3 as _sqlite3

    try:
        with _sqlite3.connect(str(DB_PATH)) as conn:
            tables = {row[0] for row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            )}
            if "employee_dispatcher_groups" not in tables:
                return

            rows = conn.execute(
                "SELECT edg.employee_id, dg.id "
                "FROM employee_dispatcher_groups edg "
                "JOIN dispatcher_groups dg ON dg.id = edg.dispatcher_group_id "
                "ORDER BY edg.employee_id, dg.name"
            ).fetchall()
            primary_by_employee = {}
            for emp_id, group_id in rows:
                primary_by_employee.setdefault(emp_id, group_id)
            for emp_id, group_id in primary_by_employee.items():
                conn.execute(
                    "UPDATE employees SET dispatcher_group_id = ? "
                    "WHERE id = ? AND dispatcher_group_id IS NULL",
                    (group_id, emp_id),
                )
            conn.commit()
            conn.execute("DROP TABLE employee_dispatcher_groups")
            conn.commit()
    except Exception as e:
        logging.error(f"Fejl ved migrering af disponentgruppe til én-til-én: {e}")


def _ensure_manage_baselines_permission():
    """Tilføjer manage_baselines til admin-rollen (idempotent)."""
    from database.models import Role
    db = SessionLocal()
    try:
        role = db.query(Role).filter(Role.name == "admin").first()
        if role:
            perms = list(role.permissions or [])
            if "manage_baselines" not in perms:
                perms.append("manage_baselines")
                role.permissions = perms
                db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved opdatering af manage_baselines-tilladelse: {e}")
    finally:
        db.close()


def _ensure_manage_auto_approval_permission():
    """Tilføjer manage_auto_approval til admin-rollen (idempotent)."""
    from database.models import Role
    db = SessionLocal()
    try:
        role = db.query(Role).filter(Role.name == "admin").first()
        if role:
            perms = list(role.permissions or [])
            if "manage_auto_approval" not in perms:
                perms.append("manage_auto_approval")
                role.permissions = perms
                db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved opdatering af manage_auto_approval-tilladelse: {e}")
    finally:
        db.close()


def _ensure_system_settings():
    """Opretter singleton-recorden for systemindstillinger hvis den mangler (idempotent)."""
    from database.models import SystemSettings
    db = SessionLocal()
    try:
        if db.query(SystemSettings).filter(SystemSettings.id == 1).first() is None:
            db.add(SystemSettings(id=1, auto_approval_enabled=True))
            db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved oprettelse af systemindstillinger: {e}")
    finally:
        db.close()


def _ensure_auto_approve_permission():
    """Tilføjer auto_approve_manual_activities til lonbogholder-rollen (idempotent)."""
    from database.models import Role
    db = SessionLocal()
    try:
        role = db.query(Role).filter(Role.name == "lonbogholder").first()
        if role:
            perms = list(role.permissions or [])
            if "auto_approve_manual_activities" not in perms:
                perms.append("auto_approve_manual_activities")
                role.permissions = perms
                db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved opdatering af auto_approve_manual_activities-tilladelse: {e}")
    finally:
        db.close()


def _ensure_activity_permissions():
    """Tilføjer approve_activities og view_calendar til alle roller (idempotent)."""
    from database.models import Role
    db = SessionLocal()
    try:
        new_perms = ["approve_activities", "view_calendar"]
        for role in db.query(Role).all():
            perms = list(role.permissions or [])
            changed = False
            for p in new_perms:
                if p not in perms:
                    perms.append(p)
                    changed = True
            if changed:
                role.permissions = perms
        db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved opdatering af aktivitetsrettigheder: {e}")
    finally:
        db.close()


def _ensure_vagtplan_permissions():
    """Tilføjer vagtplan_view + vagtplan_edit_all til ALLE roller (idempotent) – 'alle
    nuværende roller skal kunne se og redigere i vagtplanen' (spec-beslutning 2026-08-21).
    vagtplan_edit_own tilføjes IKKE automatisk – det er en mere restriktiv, opt-in ret."""
    from database.models import Role
    db = SessionLocal()
    try:
        new_perms = ["vagtplan_view", "vagtplan_edit_all"]
        for role in db.query(Role).all():
            perms = list(role.permissions or [])
            changed = False
            for p in new_perms:
                if p not in perms:
                    perms.append(p)
                    changed = True
            if changed:
                role.permissions = perms
        db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved opdatering af vagtplan-rettigheder: {e}")
    finally:
        db.close()


def _ensure_employee_supplements_permission():
    """Tilføjer manage_employee_supplements til lonbogholder-rollen (idempotent)."""
    from database.models import Role
    db = SessionLocal()
    try:
        role = db.query(Role).filter(Role.name == "lonbogholder").first()
        if role and not role.is_system:
            perms = list(role.permissions or [])
            if "manage_employee_supplements" not in perms:
                perms.append("manage_employee_supplements")
                role.permissions = perms
                db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved opdatering af manage_employee_supplements-tilladelse: {e}")
    finally:
        db.close()


def _ensure_toggle_springer_permission():
    """Tilføjer toggle_springer til ALLE roller (idempotent)."""
    from database.models import Role
    db = SessionLocal()
    try:
        for role in db.query(Role).all():
            perms = list(role.permissions or [])
            if "toggle_springer" not in perms:
                perms.append("toggle_springer")
                role.permissions = perms
        db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved opdatering af toggle_springer-tilladelse: {e}")
    finally:
        db.close()


def _ensure_payroll_settlement_permissions():
    """Tilføjer payroll_settlement_view + payroll_settlement_export til lonbogholder-rollen (idempotent)."""
    from database.models import Role
    db = SessionLocal()
    try:
        role = db.query(Role).filter(Role.name == "lonbogholder").first()
        if role and not role.is_system:
            perms = list(role.permissions or [])
            changed = False
            for p in ("payroll_settlement_view", "payroll_settlement_export"):
                if p not in perms:
                    perms.append(p)
                    changed = True
            if changed:
                role.permissions = perms
                db.commit()
    except Exception as e:
        db.rollback()
        logging.error(f"Fejl ved opdatering af payroll_settlement-tilladelser: {e}")
    finally:
        db.close()
